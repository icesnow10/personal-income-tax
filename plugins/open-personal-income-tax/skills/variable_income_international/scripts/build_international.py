#!/usr/bin/env python3
"""variable_income_international — the FOREIGN variable-income slice of the IRPF, from the informes.

Foreign stocks / ETFs (held at Avenue, Nomad, Interactive Brokers, etrade, …) are NOT on the B3, so
there is no preço médio to reconstruct and no Posição to validate against — the value comes straight
from the broker informe. This skill takes the international renda-variável items out of
processed/informes.json and writes an Excel:
  - bens_e_direitos       : the foreign equity/ETF positions (valor do informe, localização ≠ 105)
  - rendimentos_exterior  : any foreign income the informe carries — FLAGGED, because dividends/ganhos
                            no exterior go in OTHER fichas (Rendimentos Tributáveis Recebidos de PF/
                            Exterior / carnê-leão / GCAP), NOT isentos/exclusiva. For human review.

Scope: owns informe bens that are `b3:false`, foreign (localização ≠ 105) and equity-like (grupo
3/7/31). It does NOT own foreign CASH (grupo 06) or crypto (grupo 08) — those stay with /consolidate.

Usage:
  python build_international.py --informes processed/informes.json --out processed/variable_income_international.xlsx

Requires: openpyxl
"""
import argparse, json, unicodedata
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PURPLE, LPURPLE = "FF674EA7", "FFD9D2E9"
EQUITY_GRUPOS = {3, 7, 31}            # ação / fundo-ETF / outros — renda variável (não 6=cash, 8=cripto, 4=RF)


def norm(s):
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().strip()


def gint(v):
    try: return int(float(v))
    except (TypeError, ValueError): return None


def brl(x):
    try: return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError): return ""


def hdr(ws, n):
    for c in range(1, n + 1):
        cell = ws.cell(1, c)
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def sheet(wb, title, cols, rows, first=False):
    ws = wb.active if first else wb.create_sheet(title)
    ws.title = title
    ws.append(cols); hdr(ws, len(cols))
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, 1):
        vals = [len(str(ws.cell(rr, i).value or "")) for rr in range(2, ws.max_row + 1)] or [0]
        ws.column_dimensions[get_column_letter(i)].width = min(max(max(vals + [len(str(c))]) + 2, 10), 70)
    return ws


def is_foreign(loc):
    l = norm(loc)
    return l not in ("", "105", "0", "None")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--informes", required=True)
    ap.add_argument("--out", default="processed/variable_income_international.xlsx")
    a = ap.parse_args()

    data = json.loads(Path(a.informes).read_text(encoding="utf-8"))
    bens = [it for it in (data.get("bens") or [])
            if not it.get("b3") and is_foreign(it.get("localizacao")) and gint(it.get("grupo")) in EQUITY_GRUPOS]
    owned_keys = {norm(it.get("key")).upper() for it in bens}
    # foreign income the informe happened to carry under the domestic fichas (rare) — flag for review
    rend = [it for it in (data.get("isentos", []) + data.get("exclusiva", []))
            if norm(it.get("key")).upper() in owned_keys]

    wb = Workbook()
    sheet(wb, "bens_e_direitos",
          ["key", "grupo", "codigo", "localizacao", "cnpj", "descr", "valor_2024", "valor_2025", "source"],
          [{**{"localizacao": it.get("localizacao", 249)}, **it} for it in bens], first=True)
    sheet(wb, "rendimentos_exterior",
          ["codigo", "key", "descr", "valor_2025", "obs", "source"],
          [{**it, "obs": "EXTERIOR — confira a ficha (Tributáveis Exterior / carnê-leão / GCAP), NÃO isento/exclusiva"}
           for it in rend])
    Path(a.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(a.out)

    tot = sum(float(b.get("valor_2025") or 0) for b in bens)
    print(f"OK -> {a.out}")
    print(f"  bens_e_direitos (exterior)  {len(bens)} itens, total R$ {brl(round(tot,2))}")
    if not bens:
        print("  (nenhuma renda variável no exterior no informes.json — ações/ETF estrangeiros)")
    print(f"  rendimentos_exterior        {len(rend)} (revisar ficha — não vão em isento/exclusiva)")
    print("  Lembrete: dividendos/ganhos no exterior são TRIBUTÁVEIS (carnê-leão/GCAP), não entram nas 3 fichas.")


if __name__ == "__main__":
    main()
