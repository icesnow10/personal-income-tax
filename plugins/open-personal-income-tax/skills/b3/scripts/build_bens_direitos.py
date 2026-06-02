#!/usr/bin/env python3
"""Build the IRPF "Bens e Direitos" workbook from B3 data, driven by living memory files.

Inputs
  MOV.xlsx  B3 "Movimentação" export — the HISTORICAL view (all years). One sheet, B3's fixed
            8 columns: Entrada/Saída, Data, Movimentação, Produto, Instituição, Quantidade,
            Preço unitário, Valor da Operação.
  POS.xlsx  B3 "Posição" export at the LAST DAY of the fiscal year (31/12/YYYY). Multi-sheet
            (e.g. Acoes, BDR, Fundo de Investimento, Tesouro Direto); blocks may differ.

Living memory (markdown tables — the single source of truth, pasted into aux_mapping):
  mapping_memory.md     entry_movement -> action (+ logic).   [generic; bundled fallback]
  ticker_memory.md      from_ticker -> to_ticker (renames).   [taxpayer-specific]
  overrides_memory.md   ticker,date -> qty,avg_price (cost resets for mergers). [taxpayer-specific]

Usage
  python build_bens_direitos.py MOV.xlsx POS.xlsx OUT.xlsx [--memory-dir DIR] [--year YYYY]

  --memory-dir defaults to the current folder; each memory file missing there falls back to the
  bundled copy for mapping_memory.md only (ticker/overrides default to empty, with a warning).

Average price = ACQUISITION COST ÷ quantity (cost basis, what Bens e Direitos asks for), NOT
market price. See REFERENCE.md. Not tax advice.

Requires: pandas, openpyxl
"""
import argparse, json, re, shutil, unicodedata
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PURPLE, LPURPLE = "FF674EA7", "FFD9D2E9"
SKILL_DIR = Path(__file__).resolve().parents[1]            # scripts/ -> b3/
IRPF_CODE = {"Ação": (3,1), "BDR": (4,4), "FII": (7,3), "RENDA FIXA": (4,2)}  # grupo, codigo

# empty (header-only) templates seeded into the working folder so the files are visible/editable
EMPTY_TEMPLATES = {
 "ticker_memory.md": (
   "# ticker_memory — current ticker per code\n\n"
   "Explicit renames the engine can't infer (fund mergers, BDR renames, PN→ON). FII subscription\n"
   "receipts (XXXX12 / XXXX13 → XXXX11) are folded automatically — no row needed. Add yours:\n\n"
   "| from_ticker | to_ticker | note | source |\n|---|---|---|---|\n"),
 "rf_memory.md": (
   "# rf_memory — renda fixa product renames\n\n"
   "Maps a PRIOR-year B3 renda-fixa `Produto` name to its CURRENT name, so a bond whose name\n"
   "changed between years (e.g. an issuer gaining \"- EM LIQUIDACAO EXTRAJUDICIAL\") still matches\n"
   "when filling `valor_<prior>` from `--posicao-anterior`. Match is by exact Produto text.\n"
   "Add yours:\n\n"
   "| from_produto | to_produto | note | source |\n|---|---|---|---|\n"),
 "rf_value_memory.md": (
   "# rf_value_memory — Bens e Direitos value override (CRA / debêntures)\n\n"
   "For amortizing / secondary-market fixed income (CRA, CRI, debêntures), B3 alone CANNOT produce\n"
   "the Bens e Direitos value: the purchase price embeds **juros decorridos** (accrued interest paid\n"
   "to the seller) that B3 never separates, and the principal amortizes over time. The broker informe\n"
   "(BTG/NU) gives the authoritative Saldo. Pin it here per security código, with the source. CDB,\n"
   "LCI, LCA and Tesouro do NOT need this (position quantity × acquisition unit price already matches).\n\n"
   "| codigo | valor_anterior | valor_atual | note | source |\n|---|---|---|---|---|\n"),
}


def norm(s):
    if s is None: return ""
    return unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode().lower().strip()

def brl(x):  # 1234.5 -> "1.234,50"
    return f"{x:,.2f}".replace(",","X").replace(".",",").replace("X",".")

# --------------------------- living-memory loaders ---------------------------
def parse_md_tables(path):
    """All markdown tables in a file -> list of (header, [dict rows]). A non-pipe line ends a
    table, so a memory file may carry explanatory tables alongside the data table."""
    tables, header, cur = [], None, None
    for line in Path(path).read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s.startswith("|"):
            header = None; continue                      # table boundary
        cells = [c.strip() for c in s.strip("|").split("|")]
        if set("".join(cells)) <= set("-: "):            # separator row ---|---
            continue
        if header is None:
            header = cells; cur = []; tables.append((header, cur))
        else:
            cur.append(dict(zip(header, cells)))
    return tables

def table_with(path, key):
    """Rows of the (first) table that has column `key` — picks the data table, ignoring any
    explanatory tables in the same file."""
    if path is None: return []
    for header, rows in parse_md_tables(path):
        if key in header: return rows
    return []

def resolve_memory(name, memory_dir, warn):
    """Return the path to a memory file in memory_dir, SEEDING it if absent so the user can see
    and edit it: mapping_memory.md is copied from the bundled generic table; ticker/overrides get
    an empty header-only template (never illustrative rows that could be applied by mistake)."""
    p = Path(memory_dir) / name
    if p.exists(): return p
    try:
        if name == "mapping_memory.md" and (SKILL_DIR / name).exists():
            shutil.copy(SKILL_DIR / name, p)
            warn.append(f"seeded {name} into the working folder (generic B3 table — edit if needed)")
        elif name in EMPTY_TEMPLATES:
            p.write_text(EMPTY_TEMPLATES[name], encoding="utf-8")
            warn.append(f"created empty {name} in the working folder — add your rows and re-run")
        else:
            return None
        return p
    except Exception:
        return SKILL_DIR / name if (SKILL_DIR / name).exists() else None

def load_memory(memory_dir, warn):
    classification, logic, provento = {}, {}, {}
    for d in table_with(resolve_memory("mapping_memory.md", memory_dir, warn), "entry_movement"):
        mv = d.get("entry_movement")
        if not mv: continue
        classification[mv] = (d.get("credito","no_action"), d.get("debito","no_action"))
        logic[mv] = d.get("logic","")
        provento[mv] = (d.get("provento_type") or "").strip()
    renames, ren_rows = {}, []
    for d in table_with(resolve_memory("ticker_memory.md", memory_dir, warn), "from_ticker"):
        f, t = d.get("from_ticker"), d.get("to_ticker")
        if f and t: renames[f] = t
        ren_rows.append(d)
    rf_renames, rf_rows = {}, []
    for d in table_with(resolve_memory("rf_memory.md", memory_dir, warn), "from_produto"):
        f, t = d.get("from_produto"), d.get("to_produto")
        if f and t: rf_renames[f.strip()] = t.strip()
        rf_rows.append(d)
    def _num(s):
        s = (s or "").strip().replace(".","").replace(",",".")
        try: return float(s)
        except ValueError: return None
    rf_value = {}                                          # codigo -> (valor_anterior, valor_atual, source)
    for d in table_with(resolve_memory("rf_value_memory.md", memory_dir, warn), "codigo"):
        cod = (d.get("codigo") or "").strip()
        if cod: rf_value[cod] = (_num(d.get("valor_anterior")), _num(d.get("valor_atual")), d.get("source") or "")
    return classification, logic, provento, renames, ren_rows, rf_renames, rf_rows, rf_value


RECEIPT_RE = re.compile(r"^[A-Z]{4}1[23]$")               # FII subscription receipt codes
# renda fixa product "TIPO - CÓDIGO - EMISSOR" -> the security código is the ticker (e.g.
# "CDB - CDB422GUBFR - BANCO PINE S/A" -> CDB422GUBFR). The prefix must be a known fixed-income
# type followed by " - " so equity tickers (VALE3, LCAM3, CRFB3...) never match.
RF_TICKER_RE = re.compile(r"^(?:CDB|RDB|CRA|CRI|CDCA|DEB|LCI|LCA|LF|LIG|LH)\b\s*-\s*([A-Z0-9]*\d[A-Z0-9]*)")

def raw_ticker(produto):
    if not isinstance(produto, str): return None
    p = produto.strip()
    if norm(p).startswith("tesouro"): return p[:30].strip()
    m = RF_TICKER_RE.match(p)
    if m: return m.group(1)                               # renda fixa -> security código as ticker
    return p[:6].strip()

def correct_ticker(produto, renames):
    code = raw_ticker(produto)
    if code is None: return None
    if RECEIPT_RE.match(code): code = code[:4] + "11"      # fold receipt -> main
    return renames.get(code, code)

def detect_tipo(sheet_name):
    n = norm(sheet_name)
    if "bdr" in n: return "BDR"
    if "tesouro" in n or "renda fixa" in n: return "RENDA FIXA"
    if "fundo" in n: return "FII"
    if "aco" in n: return "Ação"
    return sheet_name

def is_fi_infra(prod):
    """FI-Infra incentivado (Lei 12.431): a fund whose produto name marks it as an incentivized
    infrastructure fund — IRPF grupo 07 código 10 (alíquota 0% PF), NOT the plain FII 07/03.
    Generic name heuristic (no hard-coded ticker): produto carries 'incentivad' + 'infra'."""
    n = norm(prod)
    return "incentivad" in n and "infra" in n

def tipo_label(tipo, prod):
    """Display label for the discriminação: FI-Infra funds read 'FI-INFRA', not 'FII'."""
    return "FI-INFRA" if (tipo == "FII" and is_fi_infra(prod)) else tipo

def getcol(d, *candidates):
    keys = {norm(k): k for k in d}
    for c in candidates:
        k = keys.get(norm(c))
        if k is not None and d[k] not in (None, ""): return d[k]
    return None


# B3's Movimentação export columns. Older exports have 8 columns (Entrada/Saída … Valor da
# Operação); newer ones insert "Ano" and "Ticker", giving 10. Map by HEADER NAME (not position)
# so both layouts work — positional slicing silently misaligns the new format.
MOV_COLMAP = {
    "entry_type": ["entrada/saida", "entrada / saida", "entrada/saída", " ", ""],
    "date":       ["data"],
    "entry_movement": ["movimentacao"],
    "product":    ["produto"],
    "holder":     ["instituicao"],
    "quantity":   ["quantidade"],
    "unit_price": ["preco unitario"],
    "amount":     ["valor da operacao"],
}

def read_movimentacao(mov_path):
    """Read the Movimentação sheet and rename its columns to the canonical 8 by matching header
    names, tolerating extra columns (Ano, Ticker) in newer B3 exports. Falls back to positional
    naming only if no headers match (e.g. a headerless dump)."""
    df = pd.read_excel(mov_path, sheet_name=0)
    norm_cols = {norm(c): c for c in df.columns}
    rename = {}
    for target, cands in MOV_COLMAP.items():
        for cand in cands:
            src = norm_cols.get(norm(cand))
            if src is not None and src not in rename:
                rename[src] = target
                break
    if len(rename) >= 6:                                  # headers matched — use them
        return df.rename(columns=rename)
    df.columns = list(MOV_COLMAP)[:len(df.columns)]       # fallback: positional
    return df


# =========================== 1. movements -> avg price ===========================
def build_movements(mov_path, classification, provento, renames, year):
    raw = read_movimentacao(mov_path)
    raw = raw.reset_index().rename(columns={"index":"_ord"})
    raw["date"] = pd.to_datetime(raw["date"], dayfirst=True, errors="coerce")
    raw["quantity"] = pd.to_numeric(raw["quantity"], errors="coerce").fillna(0.0)
    raw["unit_price"] = pd.to_numeric(raw["unit_price"], errors="coerce")
    raw["amount"] = pd.to_numeric(raw["amount"], errors="coerce").fillna(0.0)
    raw["raw_ticker"] = raw["product"].map(raw_ticker)
    raw["ticker"] = raw["product"].map(lambda p: correct_ticker(p, renames))

    unknown = sorted({m for m in raw["entry_movement"].dropna().unique() if m not in classification})
    def classify(r):
        cred, deb = classification.get(r["entry_movement"], ("no_action","no_action"))
        action = cred if str(r["entry_type"]).startswith("Cred") else deb
        # Transferência - Liquidação Debito on a RECEIPT code (XXXX12/13): receipt being
        # consumed in the conversion to the main cota — not a sale of the main position.
        if (action == "sale" and r["entry_movement"] == "Transferência - Liquidação"
                and RECEIPT_RE.match(r["raw_ticker"] or "")):
            return "no_action"
        return action
    raw["emt"] = raw.apply(classify, axis=1)
    raw["amount_adjusted"] = raw.apply(
        lambda r: r["amount"] if str(r["entry_type"]).startswith("Cred") else -r["amount"], axis=1)

    def deltas(r):
        e = r["emt"]
        if e == "purchase":          return ( r["quantity"],  r["amount"])
        if e == "sale":              return (-r["quantity"], -r["amount"])
        if e == "return_of_capital":
            # signed by entry_type so a Credito+Debito pair (e.g. Restituição de Capital -
            # Transferida bookkeeping) nets to zero. Single Credito reduces cost; single
            # Debito (rare reversal) adds cost back.
            cred = str(r["entry_type"]).startswith("Cred")
            return ( 0.0, -r["amount"] if cred else r["amount"])
        return (0.0, 0.0)
    raw = pd.concat([raw, raw.apply(lambda r: pd.Series(deltas(r), index=["dq","dc"]), axis=1)], axis=1)

    # custody transfers: matched in/out pairs cancel; a lone leg is a real qty change.
    # Skip receipt codes (XXXX12/13): a Transferência there is about the right, not the main cota.
    raw["tr_dq"] = 0.0
    trmask = (raw["emt"].eq("no_action")
              & raw["entry_movement"].isin(["Transferência","Transferencia"])
              & ~raw["raw_ticker"].fillna("").str.match(RECEIPT_RE.pattern))
    net = {}
    for _, r in raw[trmask].iterrows():
        k = (r["date"], r["ticker"])
        net[k] = net.get(k,0.0) + (r["quantity"] if str(r["entry_type"]).startswith("Cred") else -r["quantity"])
    seen = set()
    for idx, r in raw[trmask].iterrows():
        k = (r["date"], r["ticker"])
        if k in seen: continue
        seen.add(k)
        if abs(net.get(k,0.0)) > 1e-9: raw.at[idx,"tr_dq"] = net[k]
    raw["dq"] = raw["dq"] + raw["tr_dq"]

    # Pre-compute the snapshot anchor per ticker: the LATEST snapshot date with NO purchase/
    # sale after it. Atualização rows are ambiguous in B3 — sometimes they reassert the
    # full position (after a merger/conversion), sometimes they record a delta from a
    # specific event (a subscription receipt becoming the main cota). The "no buy/sale
    # after" rule cleanly separates them: a snapshot followed by buys/sales is treated as
    # a delta-style event (script's running calculation is more reliable); a snapshot with
    # nothing after it is the authoritative year-end position.
    cutoff_ts = pd.Timestamp(year, 12, 31)
    anchor = {}
    for tk in raw["ticker"].dropna().unique():
        sub = raw[(raw["ticker"] == tk) & (raw["date"] <= cutoff_ts)]
        snaps = sub[sub["emt"] == "snapshot"]
        if not len(snaps): continue
        latest_snap_date = snaps["date"].max()
        if len(sub[(sub["date"] > latest_snap_date) & sub["emt"].isin(["purchase","sale"])]):
            continue                                    # buys after — snapshot is delta, not position
        snap_rows = sub[(sub["date"] == latest_snap_date) & (sub["emt"] == "snapshot") & (sub["quantity"] > 0)]
        if len(snap_rows):
            anchor[tk] = (latest_snap_date,
                          float(snap_rows.groupby("holder")["quantity"].first().sum()))

    chrono = raw.sort_values(["date","_ord"], ascending=[True, True])
    qa, ca, cyc, closed = {}, {}, {}, {}
    QA, CA, AV, CY = {}, {}, {}, {}
    last_pos_avg = {}                                   # last avg seen with qty>0 per ticker
    for d, g in chrono.groupby("date", sort=True):
        for _, r in g.iterrows():
            tk = r["ticker"]; cy = cyc.get(tk,1)
            prev_q = qa.get(tk,0.0); prev_c = ca.get(tk,0.0)
            if r["emt"] == "rebase":
                # grupamento: the row carries the NEW total quantity; cost is preserved (avg
                # rises). SET the quantity, do not add.
                q = r["quantity"]; c = prev_c
            elif r["emt"] == "sale" and abs(r["amount"]) > 0.005:
                if r["quantity"] > prev_q + 1e-6:
                    # oversell — selling MORE than is held: not a cash sale but a corporate-action
                    # artifact (e.g. a fund migration redeeming the old cotas, IRDM11→IRIM11).
                    # Preserve the cost basis; the snapshot anchor restates the real quantity.
                    q = prev_q + r["dq"]; c = prev_c
                else:
                    # real sale (proceeds > 0): cost basis leaves at the CURRENT average
                    # (quantity × preço médio), NOT the sale proceeds — proceeds only matter for
                    # ganho de capital. A full exit therefore drives cost to zero. (Fração em
                    # Ativos has amount 0 and falls through below: qty drops at zero cost, avg rises.)
                    avg_now = (prev_c / prev_q) if prev_q > 1e-9 else 0.0
                    q = prev_q + r["dq"]; c = prev_c - r["quantity"] * avg_now
            else:
                q = prev_q + r["dq"]; c = prev_c + r["dc"]
            # clean full exit (qty ≈ 0) zeroes the cost; a NEGATIVE qty is an oversell transient
            # that a snapshot anchor will restate, so the cost is kept until then.
            if abs(q) < 1e-9: c = 0.0
            if abs(q) < 1e-9 and prev_q > 1e-9: closed[tk] = True
            if r["dq"] > 0 and prev_q <= 1e-9 and closed.get(tk): cy += 1; closed[tk] = False
            qa[tk], ca[tk], cyc[tk] = q, c, cy
        # After the day's deltas, apply anchor if this is the anchor date for the ticker.
        # From this row onwards qty reflects the anchored value (the snapshot row's own
        # output below and all subsequent rows).
        for tk_a, (a_date, a_qty) in anchor.items():
            if a_date == d: qa[tk_a] = a_qty
        for _, r in g.iterrows():
            tk = r["ticker"]; q = qa[tk]; c = ca[tk]
            if q > 1e-9:
                av = round(c/q, 6); last_pos_avg[tk] = av
            else:
                # qty=0 or negative — carry forward last positive avg of this ticker (sold-out
                # tickers keep their historical PM; brief transition states stay readable).
                # Tickers that never had a positive position (renda fixa, folded receipts) stay blank.
                av = last_pos_avg.get(tk)
            QA[r["_ord"]] = round(q,4); CA[r["_ord"]] = round(c,2)
            AV[r["_ord"]] = av
            CY[r["_ord"]] = cyc[tk]
    raw["provento_type"] = raw["entry_movement"].map(lambda m: provento.get(m, ""))
    raw["quantity_accumulated"] = raw["_ord"].map(QA)
    raw["amount_adjusted"] = raw["amount_adjusted"].round(2)
    raw["avg_price"] = raw["_ord"].map(AV)
    raw["custo_acumulado"] = raw["_ord"].map(CA)    # custo de aquisição acumulado (= avg × qtd)
    raw["cycle"] = raw["_ord"].map(CY)

    cut = pd.Timestamp(year,12,31)
    summary = {}
    for tk, sub in raw.groupby("ticker"):
        s = sub[sub["date"] <= cut].sort_values(["date","_ord"])
        if len(s):
            avg = s.iloc[-1]["avg_price"]
            q   = s.iloc[-1]["quantity_accumulated"]
        else:
            avg, q = None, None
        # income comes from provento_type now (action axis is decoupled from income axis)
        inc = sub[sub["provento_type"].astype(bool)]["amount_adjusted"].sum()
        summary[tk] = {"qty": (None if q is None or pd.isna(q) else float(q)),
                       "avg": (None if avg is None or pd.isna(avg) else float(avg)),
                       "income": round(float(inc),2)}
    return raw.sort_values("_ord"), summary, unknown


# =========================== 2. position blocks ===========================
def load_position(pos_path):
    wb = load_workbook(pos_path, data_only=True); blocks = []
    for sh in wb.sheetnames:
        ws = wb[sh]; tipo = detect_tipo(sh)
        headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        rows = []
        for r in range(2, ws.max_row+1):
            vals = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
            if all(v in (None,"") for v in vals): continue
            if any(str(v).strip() == "Total" for v in vals): continue
            if vals[0] in (None,""): continue
            d = {h:(v.strip() if isinstance(v,str) else v) for h,v in zip(headers,vals)}
            d["__tipo__"] = tipo; rows.append(d)
        blocks.append((tipo, headers, rows))
    return blocks

def discriminacao(d, summary):
    tipo = d["__tipo__"]; prod = str(getcol(d,"Produto") or "").strip()
    qty = getcol(d,"Quantidade"); qtyi = int(qty) if isinstance(qty,(int,float)) else qty
    inst = getcol(d,"Instituição","Instituicao") or ""
    tk = getcol(d,"Código de Negociação","Codigo de Negociacao") or prod
    if tipo == "RENDA FIXA":
        return f"APLICACAO EM {tk} NA CORRETORA {inst}"
    if tipo == "BDR":
        ident = getcol(d,"Código ISIN / Distribuição","Codigo ISIN / Distribuicao","Código ISIN") or ""; label = "ISIN"
    else:
        ident = getcol(d,"CNPJ da Empresa","CNPJ do Fundo") or ""; label = "CNPJ"
    avg = (summary.get(tk) or {}).get("avg")
    custo = f"R$ {brl(avg)}" if avg is not None else "n/d"
    return (f"{tipo_label(tipo, prod)} {tk} // {qtyi} UNIDADES // CUSTO MEDIO: {custo} // "
            f"EMPRESA: {prod} - {label} {ident} // CUSTODIA NA CORRETORA {inst}")


def discriminacao_agg(tipo, tk, ds, summary):
    """Discriminação for ONE IRPF line that merges a ticker held across several custodians:
    total units, single custo médio, and every custodian (with its quantity) listed. Falls back
    to the single-row wording when there's only one custodian."""
    prod = str(getcol(ds[0],"Produto") or "").strip()
    def qint(d):
        q = getcol(d,"Quantidade"); return int(q) if isinstance(q,(int,float)) else q
    if tipo == "RENDA FIXA":
        cods = ", ".join(dict.fromkeys(c for d in ds for c in [getcol(d,"Código","Codigo")] if c))
        insts = ", ".join(dict.fromkeys(str(getcol(d,"Instituição","Instituicao") or "") for d in ds))
        cod_part = f" ({cods})" if cods else ""
        return f"APLICACAO EM {prod}{cod_part} NA CORRETORA {insts}"
    if tipo == "BDR":
        ident = getcol(ds[0],"Código ISIN / Distribuição","Codigo ISIN / Distribuicao","Código ISIN") or ""; label = "ISIN"
    else:
        ident = getcol(ds[0],"CNPJ da Empresa","CNPJ do Fundo") or ""; label = "CNPJ"
    total_qty = sum(q for d in ds for q in [qint(d)] if isinstance(q,(int,float)))
    avg = (summary.get(tk) or {}).get("avg")
    custo = f"R$ {brl(avg)}" if avg is not None else "n/d"
    if len(ds) == 1:
        cust = f"CUSTODIA NA CORRETORA {getcol(ds[0],'Instituição','Instituicao') or ''}"
    else:
        cust = "CUSTODIA: " + "; ".join(
            f"{qint(d)} NA CORRETORA {getcol(d,'Instituição','Instituicao') or ''}" for d in ds)
    return (f"{tipo_label(tipo, prod)} {tk} // {total_qty} UNIDADES // CUSTO MEDIO: {custo} // "
            f"EMPRESA: {prod} - {label} {ident} // {cust}")


# =========================== 3. write workbook ===========================
def hdr(ws, n, fill):
    f = PatternFill("solid",fgColor=fill); ft = Font(bold=True, color=("FFFFFFFF" if fill==PURPLE else "FF000000"))
    for j in range(1,n+1):
        ws.cell(1,j).fill=f; ws.cell(1,j).font=ft; ws.cell(1,j).alignment=Alignment(horizontal="center")

def write_workbook(out_path, mov, summary, blocks, classification, logic, provento, ren_rows, year,
                   renames=None, prior_blocks=None, rf_renames=None, rf_value=None):
    rf_value = rf_value or {}
    wb = Workbook()

    ws = wb.active; ws.title = "movements_enriched"
    cols = ["entry_type","date","entry_movement","product","holder","quantity","unit_price","amount",
            "ticker","entry_movement_type","provento_type","quantity_accumulated","cycle","amount_adjusted",
            "avg_price","custo_acumulado"]
    ws.append(cols); hdr(ws, 8, PURPLE)
    for j in range(9, len(cols)+1):
        ws.cell(1,j).fill=PatternFill("solid",fgColor=LPURPLE); ws.cell(1,j).font=Font(bold=True)
        ws.cell(1,j).alignment=Alignment(horizontal="center")
    for _, r in mov.iterrows():
        ws.append([r["entry_type"], (r["date"].to_pydatetime() if pd.notna(r["date"]) else None),
                   r["entry_movement"], r["product"], r["holder"], float(r["quantity"]),
                   (None if pd.isna(r["unit_price"]) else float(r["unit_price"])), float(r["amount"]),
                   r["ticker"], r["emt"], (r["provento_type"] or None),
                   (None if pd.isna(r["quantity_accumulated"]) else float(r["quantity_accumulated"])),
                   (None if pd.isna(r["cycle"]) else int(r["cycle"])), float(r["amount_adjusted"]),
                   (None if r["avg_price"] is None or pd.isna(r["avg_price"]) else float(r["avg_price"])),
                   (None if pd.isna(r["custo_acumulado"]) else float(r["custo_acumulado"]))])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2): row[0].number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A2"

    # aux_mapping = the two living memories pasted in (movement→action | renames)
    cs = wb.create_sheet("aux_mapping")
    head = ["entry_movement","credito","debito","provento_type","logic", None,
            "from_ticker","to_ticker","rename note","rename source"]
    cs.append(head)
    for col in [1,2,3,4,5,7,8,9,10]:
        cs.cell(1,col).fill=PatternFill("solid",fgColor=LPURPLE); cs.cell(1,col).font=Font(bold=True)
    items = list(classification.items())
    for i,(mv,(cr,db)) in enumerate(items, start=2):
        cs.cell(i,1,mv); cs.cell(i,2,cr); cs.cell(i,3,db)
        cs.cell(i,4,provento.get(mv,"") or None); cs.cell(i,5,logic.get(mv,""))
    for i,d in enumerate(ren_rows, start=2):
        cs.cell(i,7,d.get("from_ticker")); cs.cell(i,8,d.get("to_ticker"))
        cs.cell(i,9,d.get("note")); cs.cell(i,10,d.get("source"))
    for c,w in zip(range(1,11),[42,14,14,18,55,3,12,12,30,34]):
        cs.column_dimensions[get_column_letter(c)].width = w

    # avg_price_summary: per-ticker position at the END OF EACH YEAR — two columns per year,
    # accumulated_quantity_<year> and avg_price_<year>, taken from the ticker's last movement row
    # on or before 31/12 of that year. Lets the cost basis be tracked year over year, not just at
    # the latest date.
    sm = wb.create_sheet("avg_price_summary")
    is_code = lambda t: isinstance(t,str) and " " not in t and len(t) <= 6
    # renda fixa tickers (códigos / Tesouro) — kept out of the equity sheets even when a movement
    # type maps them as purchase (e.g. debêntures "COMPRA/VENDA DEFINITIVA A TERMO", Tesouro Compra).
    rf_tickers = {tk for tk, p in zip(mov["ticker"], mov["product"])
                  if isinstance(p,str) and (RF_TICKER_RE.match(p) or norm(p).startswith("tesouro"))}
    main_tickers = sorted(k for k in summary
                          if summary[k]["avg"] is not None and is_code(k) and k not in rf_tickers)
    sm_years = sorted(mov["date"].dropna().dt.year.astype(int).unique().tolist())

    # ---- renda fixa: VALOR APLICADO, never the curva ----
    # Bens e Direitos for CDB/LCA/LCI/CRA/DEB is the principal: position QUANTITY × acquisition
    # UNIT PRICE. The B3 renda-fixa position export has no "Valor Aplicado" column (only MTM /
    # CURVA / FECHAMENTO, which embed accrued, untaxed yield), so the unit price comes from the
    # movimentação COMPRA/VENDA ("Valor da Operação" ÷ quantity) per security código. Using the
    # POSITION quantity makes partial redemptions fall out correctly (qty already reduced).
    # Tesouro Direto (no parseable código) keeps its own "Valor Aplicado" column.
    RF_CODE_RE = re.compile(r'^[A-Z/]+\s*-\s*([A-Z0-9]*\d[A-Z0-9]*)')   # security código (must contain a digit)
    AMORT_PREFIX = ("CRA", "CRI", "DEB")                  # amortizing / accrued-interest (juros decorridos) papers
    def _rf_kind(name):
        n = norm(name)
        if "amortiz" in n: return "amort"
        if "juros" in n:   return "juros"
        if "compra" in n:  return "compra"
        return None
    _rf_appl, _rf_qty, rf_events = {}, {}, {}             # CDB/LCA: unit price | CRA/DEB: dated events
    for _, r in mov.iterrows():
        prod = str(r["product"]); m = RF_CODE_RE.match(prod)
        if not m: continue
        cod = m.group(1)
        if prod.upper().startswith(AMORT_PREFIX):
            k = _rf_kind(r["entry_movement"])
            if k and pd.notna(r["date"]):
                rf_events.setdefault(cod, []).append((r["date"], k, float(r["amount"] or 0)))
        elif "COMPRA" in str(r["entry_movement"]).upper() and str(r["entry_type"]).startswith("Cred"):
            _rf_appl[cod] = _rf_appl.get(cod,0.0) + float(r["amount"] or 0)
            _rf_qty[cod]  = _rf_qty.get(cod,0.0) + float(r["quantity"] or 0)
    rf_unit = {c: _rf_appl[c]/_rf_qty[c] for c in _rf_appl if _rf_qty.get(c)}

    def rf_amort_calc(cod, cutoff):                        # CRA/DEB: (compra, amortização, juros) até cutoff
        ev = rf_events.get(cod, [])
        f = lambda kind: round(sum(v for dt,k,v in ev if k==kind and dt <= cutoff), 2)
        return f("compra"), f("amort"), f("juros")

    def is_amortizable(d):
        return str(getcol(d,"Produto") or "").strip().upper().startswith(AMORT_PREFIX)

    def rf_valor(d, prior=False):
        cod = getcol(d,"Código","Codigo"); qty = getcol(d,"Quantidade")
        if is_amortizable(d):
            # CRA/CRI/DEB: the purchase price embeds juros decorridos that B3 never separates, and the
            # principal amortizes — so the authoritative value is the broker informe (rf_value_memory).
            # Without an override, fall back to compra − amortização (overstates by juros decorridos).
            ov = rf_value.get(cod)
            if ov is not None and (ov[0] if prior else ov[1]) is not None:
                return round(ov[0] if prior else ov[1], 2)
            compra, amort, _ = rf_amort_calc(cod, pd.Timestamp(prev if prior else year, 12, 31))
            v = compra - amort
            return round(v,2) if abs(v) > 0.005 else None
        u = rf_unit.get(cod)
        if u is not None and isinstance(qty,(int,float)):
            return round(qty*u, 2)                         # CDB/LCA/etc: qtd × preço de aquisição
        v = getcol(d,"Valor Aplicado")                     # Tesouro Direto carries the applied value
        return round(v,2) if isinstance(v,(int,float)) else None   # else blank (no curva fallback)
    rf_no_unit = sorted({c for tipo,_,rows in blocks if tipo=="RENDA FIXA" for d in rows
                         for c in [getcol(d,"Código","Codigo")]
                         if c and c not in rf_unit and c not in rf_events and c not in rf_value
                         and not isinstance(getcol(d,"Valor Aplicado"),(int,float))})

    def ir_key(tipo, d):
        # one line per produto (renda fixa) / per ticker (equity). The displayed ticker becomes the
        # security código only when the produto holds a SINGLE código (decided in the IRPF loop).
        prod = str(getcol(d,"Produto") or "").strip()
        if tipo == "RENDA FIXA":
            return prod
        return getcol(d,"Código de Negociação","Codigo de Negociacao") or prod

    # ---- optional prior-year authoritative position (--posicao-anterior) ----
    # The B3 position file carries quantity + market value, NOT acquisition cost, so equity/FII
    # cost basis still comes from the movements (which preserves total cost through grupamentos).
    # What the prior file authoritatively adds: the real prior-year QUANTITY (fixes corporate-
    # action quantity drift like MGLU3) and renda-fixa applied value (untracked in the movements).
    prev = year - 1
    prior_q, prior_rf, prior_total_cost = {}, {}, {}   # prior_total_cost: reconstructed cost at 31/12 prev
    def _map_code(code):
        if not isinstance(code, str): return code
        c = code.strip()
        if RECEIPT_RE.match(c): c = c[:4] + "11"          # fold subscription receipt -> main cota
        return (renames or {}).get(c, c)                  # apply taxpayer renames (old code -> current)
    if prior_blocks:
        for tipo,_,rows in prior_blocks:
            for d in rows:
                if tipo == "RENDA FIXA":
                    k = ir_key(tipo, d)                         # código (CRA/DEB) ou produto (CDB/LCI/LCA/Tesouro)
                    if not is_amortizable(d):
                        k = (rf_renames or {}).get(k, k)        # produto-keyed: aplica rename de nome
                    v = rf_valor(d, prior=True)                 # qtd × preço de aquisição (valor aplicado)
                    if k and v is not None: prior_rf[k] = prior_rf.get(k,0.0) + v
                else:
                    code = getcol(d,"Código de Negociação","Codigo de Negociacao")
                    ct = _map_code(code) if code else None
                    q = getcol(d,"Quantidade")
                    if ct and isinstance(q,(int,float)): prior_q[ct] = prior_q.get(ct,0.0) + q

    head = ["ticker"]
    for y in sm_years: head += [f"accumulated_quantity_{y}", f"avg_price_{y}", f"total_{y}"]
    sm.append(head); hdr(sm, len(head), LPURPLE)
    for tk in main_tickers:
        sub = mov[mov["ticker"] == tk].sort_values(["date","_ord"])
        row = [tk]
        for y in sm_years:
            s = sub[sub["date"] <= pd.Timestamp(y,12,31)]
            if len(s):
                q = s.iloc[-1]["quantity_accumulated"]; a = s.iloc[-1]["avg_price"]
                qv = None if pd.isna(q) else float(q)
                av = None if (a is None or pd.isna(a)) else round(float(a),6)
                total_cost = (qv*av) if (qv is not None and av is not None) else None
                if y == prev and total_cost is not None: prior_total_cost[tk] = total_cost
                # prior-year override: use the authoritative quantity, keep total cost preserved,
                # re-derive avg = total ÷ qty (corrects grupamento quantity drift, e.g. MGLU3).
                if prior_blocks and y == prev and tk in prior_q and total_cost is not None and prior_q[tk]:
                    qv = prior_q[tk]; av = round(total_cost/qv, 6)
                row.append(qv); row.append(av)
                row.append(round(total_cost,2) if total_cost is not None else None)
            else:
                row += [None, None, None]
        sm.append(row)
    sm.column_dimensions["A"].width = 14
    for j in range(2, len(head)+1): sm.column_dimensions[get_column_letter(j)].width = 22
    sm.freeze_panes = "B2"

    # income — per-ticker provento breakdown, with (interest, yield, total) FOR EACH YEAR. The
    # Movimentação export is the full history; here the three metrics are repeated per calendar
    # year so income can be read/validated year by year (NaT dates excluded).
    inc = wb.create_sheet("income")
    inc_years = sorted(mov["date"].dropna().dt.year.astype(int).unique().tolist())
    head = ["ticker"]
    for y in inc_years: head += [f"interest_{y}", f"yield_{y}", f"total_{y}"]
    inc.append(head); hdr(inc, len(head), LPURPLE)
    for tk in main_tickers:
        sub = mov[(mov["ticker"] == tk) & mov["date"].notna()]
        row, has_any = [tk], False
        for y in inc_years:
            sy = sub[sub["date"].dt.year == y]
            interest = float(sy[sy["provento_type"] == "interest_on_equity"]["amount_adjusted"].sum())
            yld = float(sy[sy["provento_type"].isin(["dividend","yield","return_of_capital"])]["amount_adjusted"].sum())
            total = float(sy[sy["provento_type"].astype(bool)]["amount_adjusted"].sum())
            if interest or yld or total: has_any = True
            row += [round(interest,2) if abs(interest)>=0.005 else None,
                    round(yld,2) if abs(yld)>=0.005 else None,
                    round(total,2) if abs(total)>=0.005 else None]
        if has_any: inc.append(row)
    inc.column_dimensions["A"].width = 14
    for j in range(2, len(head)+1): inc.column_dimensions[get_column_letter(j)].width = 12
    inc.freeze_panes = "B2"

    # Amortization (return of capital) received during the fiscal year — appended to the
    # discriminação so the declaration explains the cost-basis reduction: return of capital is NOT
    # income and is NOT a taxable gain; it lowers the acquisition cost. Built from the movements
    # (traceable), equity/FII only — renda fixa amortizing papers (CRA/CRI/DEB) are documented in
    # the renda_fixa_amortizavel sheet instead. amort_text[ticker] -> clause to concatenate.
    def _cost_at(tk, cutoff):
        sub = mov[(mov["ticker"] == tk) & (mov["date"] <= cutoff)].sort_values(["date","_ord"])
        if not len(sub): return None
        q = sub.iloc[-1]["quantity_accumulated"]; a = sub.iloc[-1]["avg_price"]
        if pd.isna(q) or float(q) <= 1e-9 or a is None or pd.isna(a): return None
        return round(float(a) * float(q), 2)
    amort_text = {}
    _am = mov[(mov["emt"] == "return_of_capital") & mov["date"].notna()]
    for tk, sub in _am.groupby("ticker"):
        if tk in rf_tickers: continue
        s = sub[sub["date"].dt.year == year].sort_values(["date","_ord"])
        total = round(float(s["amount_adjusted"].sum()), 2)         # net (a Credito+Debito pair → 0)
        if total <= 0.005: continue
        evs = [(r["date"], round(float(r["amount_adjusted"]),2), str(r["holder"] or "").strip())
               for _, r in s.iterrows() if float(r["amount_adjusted"]) > 0.005]
        parc = "; ".join(f"R$ {brl(a)} EM {d.strftime('%d/%m/%Y')}"
                         + (f" NA CORRETORA {h}" if h else "") for d, a, h in evs)
        after = _cost_at(tk, pd.Timestamp(year,12,31)); before = _cost_at(tk, pd.Timestamp(prev,12,31))
        # only state "DE ... PARA ..." when amortization was the year's only cost change (so the
        # prior-year cost minus what was returned really equals the year-end cost); else omit it.
        de_para = (f", QUE PASSOU DE R$ {brl(before)} PARA R$ {brl(after)}"
                   if before is not None and after is not None and abs((before - after) - total) < 0.01 else "")
        amort_text[tk] = (f" // AMORTIZACAO (DEVOLUCAO DE CAPITAL) EM {year}: R$ {brl(total)} "
                          f"EM {len(evs)} PARCELA(S) ({parc}). POR SER DEVOLUCAO DE CAPITAL E NAO "
                          f"RENDIMENTO, O VALOR FOI ABATIDO DO CUSTO DE AQUISICAO{de_para} E NAO "
                          f"CONSTITUI GANHO TRIBUTAVEL.")

    # Custody transfers between brokers during the fiscal year — a matched out-leg (Debito at the
    # ORIGIN broker) + in-leg (Credito at the DESTINATION) of a "Transferência" on the same date,
    # to a DIFFERENT broker. Quantity and cost don't change (the net-zero pair is handled in
    # build_movements); here we only DESCRIBE the move in the discriminação. Any asset class
    # (incl. renda fixa / Tesouro); subscription-receipt transfers (XXXX12/13) are skipped.
    transfer_text = {}
    for tk, sub in mov.groupby("ticker"):
        # custody transfers apply to ANY asset class, incl. renda fixa / Tesouro (a NU->BTG move)
        trs = sub[sub["entry_movement"].isin(["Transferência","Transferencia"]) & sub["date"].notna()
                  & (sub["date"].dt.year == year)
                  & ~sub["raw_ticker"].fillna("").str.match(RECEIPT_RE.pattern)]
        events = []
        for d, g in trs.groupby("date"):
            origins = sorted({str(r["holder"]).strip() for _, r in g.iterrows()
                              if str(r["entry_type"]).startswith("Deb") and str(r["holder"]).strip()})
            dests = sorted({str(r["holder"]).strip() for _, r in g.iterrows()
                            if str(r["entry_type"]).startswith("Cred") and str(r["holder"]).strip()})
            for o in origins:
                for de in dests:
                    if o != de: events.append((d, o, de))
        if events:
            parts = "; ".join(f"EM {d.strftime('%d/%m/%Y')} DA CORRETORA {o} PARA {de}"
                              for d, o, de in events)
            transfer_text[tk] = (f" // TRANSFERENCIA DE CUSTODIA {parts} "
                                 f"(SEM ALTERACAO DE CUSTO OU QUANTIDADE).")

    # Fund incorporations / ticker conversions completed during the fiscal year — when a position's
    # cotas came from a DIFFERENT fund/code (a ticker_memory rename whose origin ROOT differs from
    # the current one) and the old code's last movement falls in `year`. The cost carries over (the
    # conversion itself isn't a taxable event); the rename note documents the event. Receipts
    # (XXXX12/13) fold to their main code, same as the engine. Same-root renames (a transient
    # receipt of the SAME fund, e.g. IRIM15→IRIM11) are skipped — that's not an incorporation.
    def _folds_to(rt, target):
        if not rt: return False
        return rt == target or (bool(RECEIPT_RE.match(rt)) and rt[:4] + "11" == target)
    rn_to_from = {}                                        # current ticker -> [(from_ticker, note)]
    for d in ren_rows:
        f, t = d.get("from_ticker"), d.get("to_ticker")
        if f and t: rn_to_from.setdefault(t, []).append((f, (d.get("note") or "").strip()))
    incorp_text = {}
    for tk, froms in rn_to_from.items():
        sub = mov[mov["ticker"] == tk]
        if not len(sub): continue
        origins = []
        for f, note in froms:
            if f[:4] == tk[:4]: continue                   # same fund family (transient receipt) — not an incorporation
            fr = sub[sub["raw_ticker"].fillna("").apply(lambda rt: _folds_to(rt, f))]
            if not len(fr) or fr["date"].dropna().empty or fr["date"].max().year != year: continue
            origins.append((f, str(fr["product"].iloc[0] or "").strip(), note))
        if origins:
            # the produto name already begins with the código, so don't repeat the bare ticker
            seg = "; ".join((prod or f) + (f" - {note}" if note else "") for f, prod, note in origins)
            incorp_text[tk] = (f" // CONVERSAO/INCORPORACAO EM {year}: POSICAO ORIGINADA DE {seg}. "
                               f"CUSTO DE AQUISICAO MANTIDO DA POSICAO ORIGINAL "
                               f"(EVENTO NAO TRIBUTAVEL NA CONVERSAO DE COTAS).")

    union = []
    for _,headers,_ in blocks:
        for h in headers:
            hh = "Tipo (B3)" if h == "Tipo" else h
            if hh not in union: union.append(hh)
    front = ["tipo","Código de Negociação","Produto","Quantidade","avg_price","custo_total"]
    rest = [h for h in union if h not in ("Código de Negociação","Produto","Quantidade")]
    pcols = front + rest + ["discriminacao"]
    ps = wb.create_sheet("position"); ps.append(pcols); hdr(ps, len(pcols), LPURPLE)
    # original position-file columns in dark purple (like movements_enriched's B3 columns); the
    # columns this script adds (tipo, avg_price, custo_total, discriminacao) stay light purple.
    derived_pos = {"tipo","avg_price","custo_total","discriminacao"}
    for j,col in enumerate(pcols,1):
        if col not in derived_pos:
            ps.cell(1,j).fill = PatternFill("solid", fgColor=PURPLE)
            ps.cell(1,j).font = Font(bold=True, color="FFFFFFFF")
            ps.cell(1,j).alignment = Alignment(horizontal="center")
    for tipo,headers,rows in blocks:
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao"); qty = getcol(d,"Quantidade")
            avg = (summary.get(tk) or {}).get("avg")
            if tipo == "RENDA FIXA":
                custo = rf_valor(d)                        # qtd × preço de aquisição (sem curva)
            else:
                custo = round(avg*qty,2) if (avg is not None and isinstance(qty,(int,float))) else None
            line = []
            for col in pcols:
                if col == "tipo": line.append(tipo)
                elif col == "avg_price": line.append(avg)
                elif col == "custo_total": line.append(custo)
                elif col == "discriminacao":
                    line.append(discriminacao(d, summary) + incorp_text.get(tk, "")
                                + amort_text.get(tk, "") + transfer_text.get(tk, ""))
                elif col == "Tipo (B3)": line.append(d.get("Tipo"))
                else: line.append(d.get(col))
            ps.append(line)
    for j,col in enumerate(pcols,1):
        ps.column_dimensions[get_column_letter(j)].width = {"Produto":48,"discriminacao":120,
            "Instituição":24,"tipo":11,"Código de Negociação":16}.get(col,15)
    ps.freeze_panes = "A2"

    # position_anterior: the prior-year B3 position as-exported (only when --posicao-anterior is
    # given), kept in the workbook for auditing. avg_price_prev / custo_total_prev are the prior-
    # year cost basis reconstructed from the movements (cost preserved across corporate actions);
    # renda fixa custo_total_prev is the applied value straight from the prior position file.
    if prior_blocks:
        punion = []
        for _,headers,_ in prior_blocks:
            for h in headers:
                hh = "Tipo (B3)" if h == "Tipo" else h
                if hh not in punion: punion.append(hh)
        pfront = ["tipo","Código de Negociação","Produto","Quantidade","avg_price_prev","custo_total_prev"]
        prest = [h for h in punion if h not in ("Código de Negociação","Produto","Quantidade")]
        pcols2 = pfront + prest
        pa = wb.create_sheet("position_anterior"); pa.append(pcols2); hdr(pa, len(pcols2), LPURPLE)
        derived_pa = {"tipo","avg_price_prev","custo_total_prev"}
        for j,col in enumerate(pcols2,1):
            if col not in derived_pa:
                pa.cell(1,j).fill = PatternFill("solid", fgColor=PURPLE)
                pa.cell(1,j).font = Font(bold=True, color="FFFFFFFF")
                pa.cell(1,j).alignment = Alignment(horizontal="center")
        for tipo,headers,rows in prior_blocks:
            for d in rows:
                code = getcol(d,"Código de Negociação","Codigo de Negociacao")
                ct = _map_code(code) if code else None
                qty = getcol(d,"Quantidade")
                if tipo == "RENDA FIXA":
                    avg_p = None
                    custo_p = rf_valor(d, prior=True)      # qtd × preço de aquisição (sem curva)
                else:
                    tc, pq = prior_total_cost.get(ct), prior_q.get(ct)
                    avg_p = round(tc/pq,6) if (tc is not None and pq) else None
                    custo_p = round(avg_p*qty,2) if (avg_p is not None and isinstance(qty,(int,float))) else None
                line = []
                for col in pcols2:
                    if col == "tipo": line.append(tipo)
                    elif col == "avg_price_prev": line.append(avg_p)
                    elif col == "custo_total_prev": line.append(custo_p)
                    elif col == "Tipo (B3)": line.append(d.get("Tipo"))
                    else: line.append(d.get(col))
                pa.append(line)
        for j,col in enumerate(pcols2,1):
            pa.column_dimensions[get_column_letter(j)].width = {"Produto":48,
                "Instituição":24,"tipo":11,"Código de Negociação":16}.get(col,15)
        pa.freeze_panes = "A2"

    # Reconciliation: posição vs movimentação por ticker (revela renames/eventos faltando)
    rc = wb.create_sheet("Reconciliation")
    rc.append(["ticker","tipo","position_qty","movement_qty_year_end","diff","status"])
    hdr(rc, 6, LPURPLE)
    cut = pd.Timestamp(year, 12, 31)
    pos_q, pos_tipo = {}, {}
    for tipo,_,rows in blocks:
        if tipo == "RENDA FIXA": continue                        # reconciliação é só de ações/FII
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao") or str(getcol(d,"Produto") or "").strip()
            q = getcol(d,"Quantidade") or 0
            if tk and isinstance(q,(int,float)):
                pos_q[tk] = pos_q.get(tk,0) + q                   # SOMA se ticker tem multiplas linhas
                pos_tipo[tk] = tipo
    mov_yq = {}
    for tk, sub in mov.groupby("ticker"):
        if tk in rf_tickers: continue                            # renda fixa não entra na reconciliação
        s = sub[sub["date"] <= cut].sort_values(["date","_ord"])
        if len(s) and pd.notna(s.iloc[-1]["quantity_accumulated"]):
            mov_yq[tk] = float(s.iloc[-1]["quantity_accumulated"])
    for tk in sorted(set(pos_q) | set(mov_yq)):
        tipo = pos_tipo.get(tk, "—")
        if tipo == "RENDA FIXA": continue
        pq, mq = pos_q.get(tk), mov_yq.get(tk)
        pq_v, mq_v = (pq or 0), (mq or 0)
        diff = round(mq_v - pq_v, 4)
        if pq is None: status = "só na movimentação (vendido ou rename faltando?)"
        elif mq is None: status = "só na posição (rename faltando?)"
        elif abs(diff) < 1e-6: status = "OK"
        else: status = "DIFERE"
        rc.append([tk, tipo, pq, mq, diff, status])
    for c,w in zip(range(1,7),[14,12,14,22,10,46]): rc.column_dimensions[get_column_letter(c)].width=w
    rc.freeze_panes = "A2"

    # Reconciliation_anterior: prior-year position (authoritative) vs the quantity reconstructed
    # from the movements at 31/12 of the prior year. Surfaces grupamentos (qty drift), renames and
    # assets the movements can't reach back to. Only built when --posicao-anterior is supplied.
    if prior_blocks:
        rca = wb.create_sheet("Reconciliation_anterior")
        rca.append([f"ticker", "position_qty_"+str(prev), "movement_qty_"+str(prev), "diff", "status"])
        hdr(rca, 5, LPURPLE)
        cutp = pd.Timestamp(prev, 12, 31)
        recon_prev = {}
        for tk, sub in mov.groupby("ticker"):
            if tk in rf_tickers: continue                        # renda fixa fora da reconciliação
            s = sub[sub["date"] <= cutp].sort_values(["date","_ord"])
            if len(s) and pd.notna(s.iloc[-1]["quantity_accumulated"]) and float(s.iloc[-1]["quantity_accumulated"]) > 1e-9:
                recon_prev[tk] = float(s.iloc[-1]["quantity_accumulated"])
        for tk in sorted(set(prior_q) | set(recon_prev)):
            pq, mq = prior_q.get(tk), recon_prev.get(tk)
            diff = round((mq or 0) - (pq or 0), 4)
            if pq is None: status = "só na movimentação (rename ou vendido?)"
            elif mq is None: status = "só na posição (custo não reconstruível das movimentações)"
            elif abs(diff) < 1e-6: status = "OK"
            else: status = "DIFERE (evento corporativo — custo total preservado, qtd ajustada)"
            rca.append([tk, pq, mq, diff, status])
        for c,w in zip(range(1,6),[14,18,18,10,54]): rca.column_dimensions[get_column_letter(c)].width=w
        rca.freeze_panes = "A2"

    ir = wb.create_sheet("IRPF_bens_e_direitos")
    # Bens e Direitos asks TWO situations: 31/12 of the prior year and 31/12 of the current
    # (fiscal) year. valor_<year> is the current cost basis (avg × year-end position); valor_
    # <year-1> is the cost basis reconstructed at 31/12 of the prior year from the movements
    # (avg × accumulated qty then). Blank for assets not yet held a year earlier. When a prior-
    # year position is supplied (--posicao-anterior), renda fixa uses its authoritative applied
    # value (else blank, since movements don't track renda-fixa cost). (prev = year-1, set above.)
    ir.append(["ticker","grupo","codigo","localizacao","cnpj","discriminacao",
               f"valor_{prev}", f"valor_{year}"]); hdr(ir,8,LPURPLE)

    def cost_at(tk, cutoff):
        sub = mov[(mov["ticker"] == tk) & (mov["date"] <= cutoff)].sort_values(["date","_ord"])
        if not len(sub): return None
        q = sub.iloc[-1]["quantity_accumulated"]; a = sub.iloc[-1]["avg_price"]
        if pd.isna(q) or float(q) <= 1e-9 or a is None or pd.isna(a): return None
        return round(float(a) * float(q), 2)

    # ONE line per ticker: a holding split across custodians (e.g. ALUP11 in BTG + XP) is merged —
    # quantities and valor summed, the custodians listed in the discriminação. Group by (tipo,
    # ticker), preserving first-seen order.
    agg, order = {}, []
    for tipo,headers,rows in blocks:
        for d in rows:
            key = (tipo, ir_key(tipo, d))
            if key not in agg: agg[key] = []; order.append(key)
            agg[key].append(d)
    for tipo, tk in order:
        ds = agg[(tipo, tk)]
        grupo, codigo = IRPF_CODE.get(tipo, ("",""))
        # isentos (LCI/LCA/CRA/CRI/debênture incentivada) são grupo 04 / código 03; tributados
        # (CDB, Tesouro/LTN, RDB) são 04 / 02. A B3 só distingue pelo prefixo do produto.
        if tipo == "RENDA FIXA":
            prod0 = str(getcol(ds[0],"Produto") or "").strip().upper()
            codigo = 3 if prod0.startswith(("LCI","LCA","CRA","CRI","DEB")) else 2
        elif tipo == "FII" and is_fi_infra(getcol(ds[0],"Produto")):
            codigo = 10                                    # FI-Infra incentivado (Lei 12.431) = 07/10, não 07/03
        cnpj = getcol(ds[0],"CNPJ da Empresa","CNPJ do Fundo") or ""
        if tipo == "RENDA FIXA":
            # VALOR APLICADO = qtd × preço de aquisição (sem curva). Tesouro usa "Valor Aplicado".
            vals = [v for v in (rf_valor(d) for d in ds) if v is not None]
            valor_cur = round(sum(vals),2) if vals else None
            # prior-year renda fixa: applied value from --posicao-anterior (by produto)
            valor_prev = round(prior_rf[tk],2) if (prior_blocks and tk in prior_rf) else None
        else:
            avg = (summary.get(tk) or {}).get("avg")
            total_qty = sum(q for d in ds for q in [getcol(d,"Quantidade")] if isinstance(q,(int,float)))
            valor_cur = round(avg*total_qty,2) if (avg is not None and total_qty) else None
            valor_prev = cost_at(tk, pd.Timestamp(prev,12,31))
        # ticker label: use the security código ONLY when the produto holds a SINGLE código (e.g.
        # ENAT14, CRA025006SS, a lone CDB). With several códigos (e.g. Banco Master), keep the
        # produto name in the ticker column and list the códigos only in the discriminação.
        ticker_label = tk
        if tipo == "RENDA FIXA":
            cods = list(dict.fromkeys(c for d in ds for c in [getcol(d,"Código","Codigo")] if c))
            if len(cods) == 1: ticker_label = cods[0]
        ir.append([ticker_label, grupo, codigo, 105, cnpj,
                   discriminacao_agg(tipo, tk, ds, summary) + incorp_text.get(tk, "")
                   + amort_text.get(tk, "") + transfer_text.get(tk, ""),
                   valor_prev, valor_cur])
    for c,w in zip(range(1,9),[12,7,7,12,18,120,14,14]): ir.column_dimensions[get_column_letter(c)].width = w
    ir.freeze_panes = "A2"

    # renda_fixa_amortizavel: the CRA/CRI/debênture calculation, shown explicitly. From B3 we get
    # compra, amortização and juros pagos; the Bens e Direitos value (saldo_declarado) is the broker
    # informe override when present (else compra − amortização). juros_decorridos is the implied
    # accrued interest paid at purchase = (compra − amort) − saldo, which B3 doesn't separate.
    cra = wb.create_sheet("renda_fixa_amortizavel")
    cra.append(["codigo","produto","corretora","valor_compra","amortizacao","juros_pagos",
                "compra_menos_amort","saldo_declarado","juros_decorridos","fonte"])
    hdr(cra, 10, LPURPLE)
    cutoff = pd.Timestamp(year,12,31)
    seen_cra = set()
    for tipo,_,rows in blocks:
        for d in rows:
            if not is_amortizable(d): continue
            cod = getcol(d,"Código","Codigo")
            if not cod or cod in seen_cra: continue
            seen_cra.add(cod)
            compra, amort, juros = rf_amort_calc(cod, cutoff)
            valor_b3 = round(compra - amort, 2)
            ov = rf_value.get(cod)
            saldo = round(ov[1],2) if (ov and ov[1] is not None) else valor_b3
            jd = round(valor_b3 - saldo, 2) if saldo is not None else None
            fonte = ov[2] if (ov and ov[1] is not None) else "B3 (compra − amortização)"
            cra.append([cod, str(getcol(d,"Produto") or "").strip(),
                        getcol(d,"Instituição","Instituicao") or "",
                        compra, amort, juros, valor_b3, saldo, jd, fonte])
    for c,w in zip(range(1,11),[14,46,24,14,13,12,18,16,16,40]):
        cra.column_dimensions[get_column_letter(c)].width = w
    cra.freeze_panes = "A2"

    # ---- IRPF rendimentos isentos (tipo 9: dividendos + rendimentos FII) and exclusivos
    # (tipo 10: JCP). Per-ticker totals, only main tickers (equity/FII/BDR), CNPJ and
    # ticker-nome resolved from the position blocks. Tickers with zero income are skipped.
    pos_lookup = {}                                        # ticker → (cnpj, full produto name)
    for tipo,_,rows in blocks:
        if tipo == "RENDA FIXA": continue
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao")
            if tk and tk not in pos_lookup:
                pos_lookup[tk] = (getcol(d,"CNPJ da Empresa","CNPJ do Fundo") or "",
                                  str(getcol(d,"Produto") or "").strip())

    # One column PER YEAR (not a single summed "valor"): the value columns are the calendar
    # years present in the history, so each ticker's provento is broken out by the year it was
    # received — the user reads the declaration year's column and validates the others. Each year
    # column is followed by an auxiliary "<year>_corretoras" column holding a JSON of the per-
    # broker (holder) split, so a year's total can be reconciled against each corretora's informe.
    def build_income_sheet(name, groups, hide_old=False, freeze=True):
        # groups = [(codigo_irpf, {provento_types}), ...] — uma mesma ficha pode ter códigos distintos
        # (ex.: isentos: dividendo=09 e rendimento de FII=99). Cada linha é marcada com o seu código.
        # hide_old: collapse (hide) year columns older than the prior year, keeping only the current
        # (year) and previous (prev) years visible. freeze: whether to freeze the first 4 columns.
        all_types = set().union(*[m for _, m in groups])
        prov = mov[mov["provento_type"].isin(all_types) & mov["date"].notna()]
        years = sorted(prov["date"].dt.year.astype(int).unique().tolist())
        ws = wb.create_sheet(name)
        head = ["tipo_de_rendimento","ticker","cnpj","ticker_nome"]
        for y in years: head += [str(y), f"{y}_corretoras"]
        ws.append(head); hdr(ws, len(head), LPURPLE)
        out = []
        for codigo, mask in groups:
            for tk, sub in mov.groupby("ticker"):
                if not is_code(tk): continue
                p = sub[sub["provento_type"].isin(mask) & sub["date"].notna()]
                if p["amount_adjusted"].abs().sum() < 0.005: continue
                p = p.assign(_y=p["date"].dt.year.astype(int))
                by_year = p.groupby("_y")["amount_adjusted"].sum()
                cnpj, produto = pos_lookup.get(tk, ("", ""))
                row = [codigo, tk, cnpj, produto]
                for y in years:
                    v = float(by_year.get(y, 0.0))
                    row.append(round(v, 2) if abs(v) >= 0.005 else None)
                    # per-broker split for this year, as compact JSON (sorted by broker for stability)
                    split = (p[p["_y"] == y].groupby("holder")["amount_adjusted"].sum().round(2).to_dict())
                    split = {str(k): val for k, val in split.items() if abs(val) >= 0.005}
                    row.append(json.dumps(dict(sorted(split.items())), ensure_ascii=False) if split else None)
                out.append((codigo, tk, row))
        for _, _, row in sorted(out, key=lambda x: (x[0], x[1])): ws.append(row)
        for c,w in zip(range(1,5),[20,12,18,60]): ws.column_dimensions[get_column_letter(c)].width = w
        for j in range(5, len(head)+1):
            col = head[j-1]
            cd = ws.column_dimensions[get_column_letter(j)]
            cd.width = 40 if col.endswith("_corretoras") else 12
            # keep only the current (year) and prior (prev) year columns visible; hide older ones
            if hide_old and int(col.split("_")[0]) < prev: cd.hidden = True
        if freeze: ws.freeze_panes = "E2"

    # Rendimentos Isentos: dividendos de ação = código 09; rendimentos de FII = código 99 ("Outros",
    # sem linha dedicada na ficha; descrição "0703 - Fundos de Investimento Imobiliário"). Ambos
    # isentos, mas códigos diferentes. Tributação Exclusiva: JCP = código 10.
    # return_of_capital (amortização) NÃO entra aqui: é devolução de capital, não rendimento — o
    # efeito já está no custo de aquisição (Bens e Direitos); declará-lo como renda seria duplicar.
    build_income_sheet("IRPF_rendimentos_isentos", [(9, {"dividend"}), (99, {"yield"})],
                       hide_old=True, freeze=False)
    build_income_sheet("IRPF_rendimentos_exclusivos", [(10, {"interest_on_equity"})],
                       hide_old=True, freeze=False)

    if rf_no_unit:
        print("NOTE: renda fixa sem COMPRA na movimentação e sem 'Valor Aplicado' na posição — "
              "valor aplicado não derivável, valor fica em branco (preencher à mão):\n  "
              + "\n  ".join(rf_no_unit))

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Build IRPF Bens e Direitos workbook from B3 data + memory files.")
    ap.add_argument("movimentacao"); ap.add_argument("posicao"); ap.add_argument("saida")
    ap.add_argument("--memory-dir", default=".", help="folder with the *_memory.md files (default: current)")
    ap.add_argument("--year", type=int)
    ap.add_argument("--posicao-anterior", dest="posicao_anterior", default=None,
                    help="B3 Posição export at 31/12 of the PRIOR year — fills the valor_<prev> column "
                         "with authoritative renda-fixa values and corrects prior-year quantities")
    a = ap.parse_args()
    warn = []
    classification, logic, provento, renames, ren_rows, rf_renames, rf_rows, rf_value = load_memory(a.memory_dir, warn)
    peek = read_movimentacao(a.movimentacao)
    year = a.year or int(pd.to_datetime(peek["date"], dayfirst=True, errors="coerce").dt.year.max())
    mov, summary, unknown = build_movements(a.movimentacao, classification, provento, renames, year)
    blocks = load_position(a.posicao)
    prior_blocks = load_position(a.posicao_anterior) if a.posicao_anterior else None
    write_workbook(a.saida, mov, summary, blocks, classification, logic, provento, ren_rows, year,
                   renames=renames, prior_blocks=prior_blocks, rf_renames=rf_renames, rf_value=rf_value)
    print(f"OK -> {a.saida}  (fiscal year {year}, {len(mov)} movement rows, "
          f"{sum(len(r) for _,_,r in blocks)} positions, {len(renames)} renames)")
    for w in warn: print("NOTE:", w)
    # prior-year renda-fixa products that did NOT match a current product name (matched by produto):
    # the valor_<prev> is left blank — fill it by hand (resgatado/vencido, or the issuer was renamed,
    # e.g. a CDB whose name gained "- EM LIQUIDACAO EXTRAJUDICIAL").
    if prior_blocks:
        def rf_products(bl):
            s = {}
            for tipo,_,rows in bl:
                if tipo != "RENDA FIXA": continue
                for d in rows:
                    prod = str(getcol(d,"Produto") or "").strip()
                    v = getcol(d,"Valor Aplicado","Valor Atualizado CURVA","Valor Atualizado MTM","Valor Atualizado")
                    if prod: s[prod] = s.get(prod,0.0) + (v if isinstance(v,(int,float)) else 0.0)
            return s
        cur_rf, prev_rf_raw = rf_products(blocks), rf_products(prior_blocks)
        prev_rf = {}                                       # apply rf_memory renames (prior -> current)
        for k, v in prev_rf_raw.items():
            kk = (rf_renames or {}).get(k, k); prev_rf[kk] = prev_rf.get(kk, 0.0) + v
        unmatched = sorted(p for p in prev_rf if p not in cur_rf)
        if unmatched:
            print(f"\nNOTE: {len(unmatched)} renda-fixa product(s) held at 31/12/{year-1} have no "
                  f"matching name in the current position - valor_{year-1} left blank, fill by hand:")
            for p in unmatched:
                print(f"    {p}  (valor {year-1} ~ R$ {brl(round(prev_rf[p],2))})")
    if unknown:
        print("WARNING: unmapped entry_movement (treated as no_action) — add to mapping_memory.md:\n  "
              + "\n  ".join(unknown))
    # ------------------------- AUDIT: movements vs position -------------------------
    pos_q, pos_tipo = {}, {}
    for tipo,_,rows in blocks:
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao")
            q  = getcol(d,"Quantidade") or 0
            if tk and isinstance(q,(int,float)) and tipo != "RENDA FIXA":
                pos_q[tk] = pos_q.get(tk,0) + q
                pos_tipo[tk] = tipo
    mismatches = []
    for tk, pq in sorted(pos_q.items()):
        mq = (summary.get(tk) or {}).get("qty") or 0
        if abs(mq - pq) > 1e-6:
            mismatches.append((tk, pos_tipo[tk], pq, mq, round(mq - pq, 4)))
    ok = len(pos_q) - len(mismatches)
    print(f"\nAUDIT (quantity in movements vs position): {ok}/{len(pos_q)} OK")
    if mismatches:
        print("  Mismatches need attention before declaring in IRPF:")
        for tk,tp,pq,mq,d in mismatches:
            print(f"    {tk:8} ({tp:5}) position={pq}  movements={mq}  diff={d:+}")
    else:
        print("  All main-ticker quantities match the year-end position.")
    # Second axis: avg_price in summary must equal the latest avg_price per ticker in the
    # movements sheet (they're computed from the same state, but a divergence would mean a
    # bug in propagation).
    avg_mismatches = []
    for tk, sub in mov.groupby("ticker"):
        s = sub[sub["date"] <= pd.Timestamp(year, 12, 31)].sort_values(["date","_ord"])
        if not len(s): continue
        mov_avg = s.iloc[-1]["avg_price"]
        sum_avg = (summary.get(tk) or {}).get("avg")
        ma = None if mov_avg is None or pd.isna(mov_avg) else round(float(mov_avg), 6)
        sa = None if sum_avg is None else round(float(sum_avg), 6)
        if ma != sa:
            avg_mismatches.append((tk, ma, sa))
    print(f"AUDIT (latest avg_price movements vs avg_price_summary): {len(summary)-len(avg_mismatches)}/{len(summary)} OK")
    for tk, ma, sa in avg_mismatches:
        print(f"    {tk:8} movements={ma}  summary={sa}")


if __name__ == "__main__":
    main()
