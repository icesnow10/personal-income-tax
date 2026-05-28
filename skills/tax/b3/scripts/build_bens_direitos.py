#!/usr/bin/env python3
"""Build the IRPF "Bens e Direitos" workbook from B3 data, driven by living memory files.

Inputs
  MOV.xlsx  B3 "Movimentação" export — the HISTORICAL view (all years). One sheet, B3's fixed
            8 columns: Entrada/Saída, Data, Movimentação, Produto, Instituição, Quantidade,
            Preço unitário, Valor da Operação.
  POS.xlsx  B3 "Posição" export at the LAST DAY of the fiscal year (31/12/YYYY). Multi-sheet
            (e.g. Acoes, BDR, Fundo de Investimento, Tesouro Direto); blocks may differ.

Living memory (markdown tables — the single source of truth, pasted into aux_mapping):
  mapping_memory.md     entry_movement -> action (+ logic).   [generic; bundled fallback]
  ticker_memory.md      from_ticker -> to_ticker (renames).   [taxpayer-specific]
  overrides_memory.md   ticker,date -> qty,avg_price (cost resets for mergers). [taxpayer-specific]

Usage
  python build_bens_direitos.py MOV.xlsx POS.xlsx OUT.xlsx [--memory-dir DIR] [--year YYYY]

  --memory-dir defaults to the current folder; each memory file missing there falls back to the
  bundled copy for mapping_memory.md only (ticker/overrides default to empty, with a warning).

Average price = ACQUISITION COST ÷ quantity (cost basis, what Bens e Direitos asks for), NOT
market price. See REFERENCE.md. Not tax advice.

Requires: pandas, openpyxl
"""
import argparse, re, shutil, unicodedata
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PURPLE, LPURPLE = "FF674EA7", "FFD9D2E9"
SKILL_DIR = Path(__file__).resolve().parents[1]            # scripts/ -> b3/
IRPF_CODE = {"Ação": (3,1), "BDR": (4,4), "FII": (7,3), "RENDA FIXA": (4,2)}  # grupo, codigo

# empty (header-only) templates seeded into the working folder so the files are visible/editable
EMPTY_TEMPLATES = {
 "ticker_memory.md": (
   "# ticker_memory — current ticker per code\n\n"
   "Explicit renames the engine can't infer (fund mergers, BDR renames, PN→ON). FII subscription\n"
   "receipts (XXXX12 / XXXX13 → XXXX11) are folded automatically — no row needed. Add yours:\n\n"
   "| from_ticker | to_ticker | note | source |\n|---|---|---|---|\n"),
}


def norm(s):
    if s is None: return ""
    return unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode().lower().strip()

def brl(x):  # 1234.5 -> "1.234,50"
    return f"{x:,.2f}".replace(",","X").replace(".",",").replace("X",".")

# --------------------------- living-memory loaders ---------------------------
def parse_md_tables(path):
    """All markdown tables in a file -> list of (header, [dict rows]). A non-pipe line ends a
    table, so a memory file may carry explanatory tables alongside the data table."""
    tables, header, cur = [], None, None
    for line in Path(path).read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s.startswith("|"):
            header = None; continue                      # table boundary
        cells = [c.strip() for c in s.strip("|").split("|")]
        if set("".join(cells)) <= set("-: "):            # separator row ---|---
            continue
        if header is None:
            header = cells; cur = []; tables.append((header, cur))
        else:
            cur.append(dict(zip(header, cells)))
    return tables

def table_with(path, key):
    """Rows of the (first) table that has column `key` — picks the data table, ignoring any
    explanatory tables in the same file."""
    if path is None: return []
    for header, rows in parse_md_tables(path):
        if key in header: return rows
    return []

def resolve_memory(name, memory_dir, warn):
    """Return the path to a memory file in memory_dir, SEEDING it if absent so the user can see
    and edit it: mapping_memory.md is copied from the bundled generic table; ticker/overrides get
    an empty header-only template (never illustrative rows that could be applied by mistake)."""
    p = Path(memory_dir) / name
    if p.exists(): return p
    try:
        if name == "mapping_memory.md" and (SKILL_DIR / name).exists():
            shutil.copy(SKILL_DIR / name, p)
            warn.append(f"seeded {name} into the working folder (generic B3 table — edit if needed)")
        elif name in EMPTY_TEMPLATES:
            p.write_text(EMPTY_TEMPLATES[name], encoding="utf-8")
            warn.append(f"created empty {name} in the working folder — add your rows and re-run")
        else:
            return None
        return p
    except Exception:
        return SKILL_DIR / name if (SKILL_DIR / name).exists() else None

def load_memory(memory_dir, warn):
    classification, logic, provento = {}, {}, {}
    for d in table_with(resolve_memory("mapping_memory.md", memory_dir, warn), "entry_movement"):
        mv = d.get("entry_movement")
        if not mv: continue
        classification[mv] = (d.get("credito","no_action"), d.get("debito","no_action"))
        logic[mv] = d.get("logic","")
        provento[mv] = (d.get("provento_type") or "").strip()
    renames, ren_rows = {}, []
    for d in table_with(resolve_memory("ticker_memory.md", memory_dir, warn), "from_ticker"):
        f, t = d.get("from_ticker"), d.get("to_ticker")
        if f and t: renames[f] = t
        ren_rows.append(d)
    return classification, logic, provento, renames, ren_rows


RECEIPT_RE = re.compile(r"^[A-Z]{4}1[23]$")               # FII subscription receipt codes

def raw_ticker(produto):
    if not isinstance(produto, str): return None
    p = produto.strip()
    return p[:30].strip() if norm(p).startswith("tesouro") else p[:6].strip()

def correct_ticker(produto, renames):
    code = raw_ticker(produto)
    if code is None: return None
    if RECEIPT_RE.match(code): code = code[:4] + "11"      # fold receipt -> main
    return renames.get(code, code)

def detect_tipo(sheet_name):
    n = norm(sheet_name)
    if "bdr" in n: return "BDR"
    if "tesouro" in n or "renda fixa" in n: return "RENDA FIXA"
    if "fundo" in n: return "FII"
    if "aco" in n: return "Ação"
    return sheet_name

def getcol(d, *candidates):
    keys = {norm(k): k for k in d}
    for c in candidates:
        k = keys.get(norm(c))
        if k is not None and d[k] not in (None, ""): return d[k]
    return None


# =========================== 1. movements -> avg price ===========================
def build_movements(mov_path, classification, provento, renames, year):
    raw = pd.read_excel(mov_path, sheet_name=0)
    raw.columns = ["entry_type","date","entry_movement","product","holder",
                   "quantity","unit_price","amount"][:len(raw.columns)]
    raw = raw.reset_index().rename(columns={"index":"_ord"})
    raw["date"] = pd.to_datetime(raw["date"], dayfirst=True, errors="coerce")
    raw["quantity"] = pd.to_numeric(raw["quantity"], errors="coerce").fillna(0.0)
    raw["unit_price"] = pd.to_numeric(raw["unit_price"], errors="coerce")
    raw["amount"] = pd.to_numeric(raw["amount"], errors="coerce").fillna(0.0)
    raw["raw_ticker"] = raw["product"].map(raw_ticker)
    raw["ticker"] = raw["product"].map(lambda p: correct_ticker(p, renames))

    unknown = sorted({m for m in raw["entry_movement"].dropna().unique() if m not in classification})
    def classify(r):
        cred, deb = classification.get(r["entry_movement"], ("no_action","no_action"))
        action = cred if str(r["entry_type"]).startswith("Cred") else deb
        # Transferência - Liquidação Debito on a RECEIPT code (XXXX12/13): receipt being
        # consumed in the conversion to the main cota — not a sale of the main position.
        if (action == "sale" and r["entry_movement"] == "Transferência - Liquidação"
                and RECEIPT_RE.match(r["raw_ticker"] or "")):
            return "no_action"
        return action
    raw["emt"] = raw.apply(classify, axis=1)
    raw["amount_adjusted"] = raw.apply(
        lambda r: r["amount"] if str(r["entry_type"]).startswith("Cred") else -r["amount"], axis=1)

    def deltas(r):
        e = r["emt"]
        if e == "purchase":          return ( r["quantity"],  r["amount"])
        if e == "sale":              return (-r["quantity"], -r["amount"])
        if e == "return_of_capital":
            # signed by entry_type so a Credito+Debito pair (e.g. Restituição de Capital -
            # Transferida bookkeeping) nets to zero. Single Credito reduces cost; single
            # Debito (rare reversal) adds cost back.
            cred = str(r["entry_type"]).startswith("Cred")
            return ( 0.0, -r["amount"] if cred else r["amount"])
        return (0.0, 0.0)
    raw = pd.concat([raw, raw.apply(lambda r: pd.Series(deltas(r), index=["dq","dc"]), axis=1)], axis=1)

    # custody transfers: matched in/out pairs cancel; a lone leg is a real qty change.
    # Skip receipt codes (XXXX12/13): a Transferência there is about the right, not the main cota.
    raw["tr_dq"] = 0.0
    trmask = (raw["emt"].eq("no_action")
              & raw["entry_movement"].isin(["Transferência","Transferencia"])
              & ~raw["raw_ticker"].fillna("").str.match(RECEIPT_RE.pattern))
    net = {}
    for _, r in raw[trmask].iterrows():
        k = (r["date"], r["ticker"])
        net[k] = net.get(k,0.0) + (r["quantity"] if str(r["entry_type"]).startswith("Cred") else -r["quantity"])
    seen = set()
    for idx, r in raw[trmask].iterrows():
        k = (r["date"], r["ticker"])
        if k in seen: continue
        seen.add(k)
        if abs(net.get(k,0.0)) > 1e-9: raw.at[idx,"tr_dq"] = net[k]
    raw["dq"] = raw["dq"] + raw["tr_dq"]

    chrono = raw.sort_values(["date","_ord"], ascending=[True, False])
    qa, ca, cyc, closed = {}, {}, {}, {}
    QA, CA, AV, CY = {}, {}, {}, {}
    for d, g in chrono.groupby("date", sort=True):
        for _, r in g.iterrows():
            tk = r["ticker"]; cy = cyc.get(tk,1)
            q = qa.get(tk,0.0) + r["dq"]; c = ca.get(tk,0.0) + r["dc"]
            if abs(q) < 1e-9 and qa.get(tk,0.0) > 1e-9: closed[tk] = True
            if r["dq"] > 0 and qa.get(tk,0.0) <= 1e-9 and closed.get(tk): cy += 1; closed[tk] = False
            qa[tk], ca[tk], cyc[tk] = q, c, cy
        for _, r in g.iterrows():
            tk = r["ticker"]; q = qa[tk]; c = ca[tk]
            QA[r["_ord"]] = round(q,4); CA[r["_ord"]] = round(c,2)
            AV[r["_ord"]] = round(c/q,6) if abs(q) > 1e-9 else None
            CY[r["_ord"]] = cyc[tk]
    raw["quantity_accumulated"] = raw["_ord"].map(QA)
    raw["amount_adjusted"] = raw["amount_adjusted"].round(2)
    raw["avg_price"] = raw["_ord"].map(AV)
    raw["cycle"] = raw["_ord"].map(CY)
    raw["provento_type"] = raw["entry_movement"].map(lambda m: provento.get(m, ""))

    cut = pd.Timestamp(year,12,31)
    summary = {}
    for tk, sub in raw.groupby("ticker"):
        s = sub[sub["date"] <= cut].sort_values(["date","_ord"])
        if len(s):
            avg = s.iloc[-1]["avg_price"]
            q   = s.iloc[-1]["quantity_accumulated"]
        else:
            avg, q = None, None
        # income comes from provento_type now (action axis is decoupled from income axis)
        inc = sub[sub["provento_type"].astype(bool)]["amount_adjusted"].sum()
        summary[tk] = {"qty": (None if q is None or pd.isna(q) else float(q)),
                       "avg": (None if avg is None or pd.isna(avg) else float(avg)),
                       "income": round(float(inc),2)}
    return raw.sort_values("_ord"), summary, unknown


# =========================== 2. position blocks ===========================
def load_position(pos_path):
    wb = load_workbook(pos_path, data_only=True); blocks = []
    for sh in wb.sheetnames:
        ws = wb[sh]; tipo = detect_tipo(sh)
        headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        rows = []
        for r in range(2, ws.max_row+1):
            vals = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
            if all(v in (None,"") for v in vals): continue
            if any(str(v).strip() == "Total" for v in vals): continue
            if vals[0] in (None,""): continue
            d = {h:(v.strip() if isinstance(v,str) else v) for h,v in zip(headers,vals)}
            d["__tipo__"] = tipo; rows.append(d)
        blocks.append((tipo, headers, rows))
    return blocks

def discriminacao(d, summary):
    tipo = d["__tipo__"]; prod = str(getcol(d,"Produto") or "").strip()
    qty = getcol(d,"Quantidade"); qtyi = int(qty) if isinstance(qty,(int,float)) else qty
    inst = getcol(d,"Instituição","Instituicao") or ""
    tk = getcol(d,"Código de Negociação","Codigo de Negociacao") or prod
    if tipo == "RENDA FIXA":
        return f"APLICACAO EM {tk} NA CORRETORA {inst}"
    if tipo == "BDR":
        ident = getcol(d,"Código ISIN / Distribuição","Codigo ISIN / Distribuicao","Código ISIN") or ""; label = "ISIN"
    else:
        ident = getcol(d,"CNPJ da Empresa","CNPJ do Fundo") or ""; label = "CNPJ"
    avg = (summary.get(tk) or {}).get("avg")
    custo = f"R$ {brl(avg)}" if avg is not None else "n/d"
    return (f"{tipo} {tk} // {qtyi} UNIDADES // CUSTO MEDIO: {custo} // "
            f"EMPRESA: {prod} - {label} {ident} // CUSTODIA NA CORRETORA {inst}")


# =========================== 3. write workbook ===========================
def hdr(ws, n, fill):
    f = PatternFill("solid",fgColor=fill); ft = Font(bold=True, color=("FFFFFFFF" if fill==PURPLE else "FF000000"))
    for j in range(1,n+1):
        ws.cell(1,j).fill=f; ws.cell(1,j).font=ft; ws.cell(1,j).alignment=Alignment(horizontal="center")

def write_workbook(out_path, mov, summary, blocks, classification, logic, provento, ren_rows, year):
    wb = Workbook()

    ws = wb.active; ws.title = "movements_to_avg_price"
    cols = ["entry_type","date","entry_movement","product","holder","quantity","unit_price","amount",
            "ticker","entry_movement_type","provento_type","quantity_accumulated","cycle","amount_adjusted","avg_price"]
    ws.append(cols); hdr(ws, 8, PURPLE)
    for j in range(9, len(cols)+1):
        ws.cell(1,j).fill=PatternFill("solid",fgColor=LPURPLE); ws.cell(1,j).font=Font(bold=True)
        ws.cell(1,j).alignment=Alignment(horizontal="center")
    for _, r in mov.iterrows():
        ws.append([r["entry_type"], (r["date"].to_pydatetime() if pd.notna(r["date"]) else None),
                   r["entry_movement"], r["product"], r["holder"], float(r["quantity"]),
                   (None if pd.isna(r["unit_price"]) else float(r["unit_price"])), float(r["amount"]),
                   r["ticker"], r["emt"], (r["provento_type"] or None),
                   (None if pd.isna(r["quantity_accumulated"]) else float(r["quantity_accumulated"])),
                   (None if pd.isna(r["cycle"]) else int(r["cycle"])), float(r["amount_adjusted"]),
                   (None if r["avg_price"] is None or pd.isna(r["avg_price"]) else float(r["avg_price"]))])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2): row[0].number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A2"

    # aux_mapping = the two living memories pasted in (movement→action | renames)
    cs = wb.create_sheet("aux_mapping")
    head = ["entry_movement","credito","debito","provento_type","logic", None,
            "from_ticker","to_ticker","rename note","rename source"]
    cs.append(head)
    for col in [1,2,3,4,5,7,8,9,10]:
        cs.cell(1,col).fill=PatternFill("solid",fgColor=LPURPLE); cs.cell(1,col).font=Font(bold=True)
    items = list(classification.items())
    for i,(mv,(cr,db)) in enumerate(items, start=2):
        cs.cell(i,1,mv); cs.cell(i,2,cr); cs.cell(i,3,db)
        cs.cell(i,4,provento.get(mv,"") or None); cs.cell(i,5,logic.get(mv,""))
    for i,d in enumerate(ren_rows, start=2):
        cs.cell(i,7,d.get("from_ticker")); cs.cell(i,8,d.get("to_ticker"))
        cs.cell(i,9,d.get("note")); cs.cell(i,10,d.get("source"))
    for c,w in zip(range(1,11),[42,14,14,18,55,3,12,12,30,34]):
        cs.column_dimensions[get_column_letter(c)].width = w

    sm = wb.create_sheet("avg_price_summary")
    sm.append(["ticker","latest_accumulated_quantity","latest_avg_price","income_received"]); hdr(sm,4,LPURPLE)
    is_code = lambda t: isinstance(t,str) and " " not in t and len(t) <= 6
    for tk in sorted(k for k in summary if summary[k]["avg"] is not None and is_code(k)):
        s = summary[tk]
        sm.append([tk, s["qty"], round(s["avg"],6), s["income"]])
    for c,w in zip(range(1,5),[14,26,18,16]): sm.column_dimensions[get_column_letter(c)].width = w

    union = []
    for _,headers,_ in blocks:
        for h in headers:
            hh = "Tipo (B3)" if h == "Tipo" else h
            if hh not in union: union.append(hh)
    front = ["tipo","Código de Negociação","Produto","Quantidade","avg_price","custo_total"]
    rest = [h for h in union if h not in ("Código de Negociação","Produto","Quantidade")]
    pcols = front + rest + ["discriminacao"]
    ps = wb.create_sheet("Position"); ps.append(pcols); hdr(ps, len(pcols), LPURPLE)
    for tipo,headers,rows in blocks:
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao"); qty = getcol(d,"Quantidade")
            avg = (summary.get(tk) or {}).get("avg")
            custo = round(avg*qty,2) if (avg is not None and isinstance(qty,(int,float))) else None
            line = []
            for col in pcols:
                if col == "tipo": line.append(tipo)
                elif col == "avg_price": line.append(avg)
                elif col == "custo_total": line.append(custo)
                elif col == "discriminacao": line.append(discriminacao(d, summary))
                elif col == "Tipo (B3)": line.append(d.get("Tipo"))
                else: line.append(d.get(col))
            ps.append(line)
    for j,col in enumerate(pcols,1):
        ps.column_dimensions[get_column_letter(j)].width = {"Produto":48,"discriminacao":120,
            "Instituição":24,"tipo":11,"Código de Negociação":16}.get(col,15)
    ps.freeze_panes = "A2"

    # Reconciliation: posição vs movimentação por ticker (revela renames/eventos faltando)
    rc = wb.create_sheet("Reconciliation")
    rc.append(["ticker","tipo","position_qty","movement_qty_year_end","diff","status"])
    hdr(rc, 6, LPURPLE)
    cut = pd.Timestamp(year, 12, 31)
    pos_q, pos_tipo = {}, {}
    for tipo,_,rows in blocks:
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao") or str(getcol(d,"Produto") or "").strip()
            q = getcol(d,"Quantidade") or 0
            if tk and isinstance(q,(int,float)):
                pos_q[tk] = pos_q.get(tk,0) + q                   # SOMA se ticker tem multiplas linhas
                pos_tipo[tk] = tipo
    mov_yq = {}
    for tk, sub in mov.groupby("ticker"):
        s = sub[sub["date"] <= cut].sort_values(["date","_ord"])
        if len(s) and pd.notna(s.iloc[-1]["quantity_accumulated"]):
            mov_yq[tk] = float(s.iloc[-1]["quantity_accumulated"])
    for tk in sorted(set(pos_q) | set(mov_yq)):
        tipo = pos_tipo.get(tk, "—")
        if tipo == "RENDA FIXA": continue
        pq, mq = pos_q.get(tk), mov_yq.get(tk)
        pq_v, mq_v = (pq or 0), (mq or 0)
        diff = round(mq_v - pq_v, 4)
        if pq is None: status = "só na movimentação (vendido ou rename faltando?)"
        elif mq is None: status = "só na posição (rename faltando?)"
        elif abs(diff) < 1e-6: status = "OK"
        else: status = "DIFERE"
        rc.append([tk, tipo, pq, mq, diff, status])
    for c,w in zip(range(1,7),[14,12,14,22,10,46]): rc.column_dimensions[get_column_letter(c)].width=w
    rc.freeze_panes = "A2"

    ir = wb.create_sheet("IRPF")
    ir.append(["ticker","grupo","codigo","localizacao","cnpj","discriminacao","valor"]); hdr(ir,7,LPURPLE)
    for tipo,headers,rows in blocks:
        grupo, codigo = IRPF_CODE.get(tipo, ("",""))
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao") or str(getcol(d,"Produto") or "").strip()
            cnpj = getcol(d,"CNPJ da Empresa","CNPJ do Fundo") or ""; qty = getcol(d,"Quantidade")
            if tipo == "RENDA FIXA":
                # Tesouro: "Valor Aplicado". Renda Fixa privada (CDB/CRA/LCA/LCI/DEB): the B3
                # export carries "Valor Atualizado CURVA" (curva-based marked value).
                valor = getcol(d, "Valor Aplicado", "Valor Atualizado CURVA",
                               "Valor Atualizado MTM", "Valor Atualizado")
            else:
                avg = (summary.get(getcol(d,"Código de Negociação","Codigo de Negociacao")) or {}).get("avg")
                valor = round(avg*qty,2) if (avg is not None and isinstance(qty,(int,float))) else None
            ir.append([tk, grupo, codigo, 105, cnpj, discriminacao(d, summary), valor])
    for c,w in zip(range(1,8),[12,7,7,12,18,120,14]): ir.column_dimensions[get_column_letter(c)].width = w
    ir.freeze_panes = "A2"

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Build IRPF Bens e Direitos workbook from B3 data + memory files.")
    ap.add_argument("movimentacao"); ap.add_argument("posicao"); ap.add_argument("saida")
    ap.add_argument("--memory-dir", default=".", help="folder with the *_memory.md files (default: current)")
    ap.add_argument("--year", type=int)
    a = ap.parse_args()
    warn = []
    classification, logic, provento, renames, ren_rows = load_memory(a.memory_dir, warn)
    peek = pd.read_excel(a.movimentacao, sheet_name=0)
    peek.columns = ["entry_type","date","entry_movement","product","holder",
                    "quantity","unit_price","amount"][:len(peek.columns)]
    year = a.year or int(pd.to_datetime(peek["date"], dayfirst=True, errors="coerce").dt.year.max())
    mov, summary, unknown = build_movements(a.movimentacao, classification, provento, renames, year)
    blocks = load_position(a.posicao)
    write_workbook(a.saida, mov, summary, blocks, classification, logic, provento, ren_rows, year)
    print(f"OK -> {a.saida}  (fiscal year {year}, {len(mov)} movement rows, "
          f"{sum(len(r) for _,_,r in blocks)} positions, {len(renames)} renames)")
    for w in warn: print("NOTE:", w)
    if unknown:
        print("WARNING: unmapped entry_movement (treated as no_action) — add to mapping_memory.md:\n  "
              + "\n  ".join(unknown))
    # ------------------------- AUDIT: movements vs position -------------------------
    pos_q, pos_tipo = {}, {}
    for tipo,_,rows in blocks:
        for d in rows:
            tk = getcol(d,"Código de Negociação","Codigo de Negociacao")
            q  = getcol(d,"Quantidade") or 0
            if tk and isinstance(q,(int,float)) and tipo != "RENDA FIXA":
                pos_q[tk] = pos_q.get(tk,0) + q
                pos_tipo[tk] = tipo
    mismatches = []
    for tk, pq in sorted(pos_q.items()):
        mq = (summary.get(tk) or {}).get("qty") or 0
        if abs(mq - pq) > 1e-6:
            mismatches.append((tk, pos_tipo[tk], pq, mq, round(mq - pq, 4)))
    ok = len(pos_q) - len(mismatches)
    print(f"\nAUDIT (quantity in movements vs position): {ok}/{len(pos_q)} OK")
    if mismatches:
        print("  Mismatches need attention before declaring in IRPF:")
        for tk,tp,pq,mq,d in mismatches:
            print(f"    {tk:8} ({tp:5}) position={pq}  movements={mq}  diff={d:+}")
    else:
        print("  All main-ticker quantities match the year-end position.")


if __name__ == "__main__":
    main()
