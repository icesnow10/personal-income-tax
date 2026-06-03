#!/usr/bin/env python3
"""renda_fixa — build the IRPF renda-fixa slice from the informes + a B3 position validation.

B3 renda fixa (CDB / CRA / CRI / debênture / LCI / LCA / Tesouro) does NOT have a reconstructable
preço médio — the value comes from the broker informe. This skill takes:
  - processed/informes.json  (what /read transcribed — the RF items: value + rendimentos)
  - the B3 Posição export     (to know which RF securities exist + their quantities → validation)
and writes an Excel with 4 sheets, RF only:
  - position             : the B3 renda-fixa position (código, produto, quantidade, corretora) + a
                           `validacao` column = whether the security is covered by an informe item
  - bens_e_direitos      : RF bens (grupo 04/02 tributado, 04/03 isento) — value from the informe
  - isentos              : RF rendimentos isentos — código 12 (juros de CRA/CRI/LCI/LCA/deb. infra)
  - exclusiva            : RF rendimentos exclusiva — código 06 (CDB/Tesouro/aplicações)

Ownership: this skill owns the B3 renda-fixa items (those whose key/produto matches the B3 Posição).
Bank products NOT in the B3 position (RDB, conta, come-cotas, exterior) are NOT renda_fixa — they
stay with /consolidate. The quantity check (position B3 × informe) only WARNS; it never blocks.

Usage:
  python build_fixed_income.py --posicao resources/POS.xlsx --informes processed/informes.json \
      --out processed/fixed_income.xlsx

Requires: pandas, openpyxl
"""
import argparse, json, re, unicodedata
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PURPLE, LPURPLE = "FF674EA7", "FFD9D2E9"
RF_PREFIX = ("CDB", "RDB", "CRA", "CRI", "CDCA", "DEB", "LCI", "LCA", "LF", "LIG", "LH", "LCD")


def norm(s):
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().strip()


def getcol(d, *names):
    for n in names:
        for k in d:
            if norm(k).lower() == norm(n).lower():
                return d[k]
    return None


def read_rf_position(path):
    """Read the B3 Posição renda-fixa / Tesouro sheets → [{codigo, produto, quantidade, corretora}]."""
    if not path or not Path(path).exists():
        return []
    xl = pd.ExcelFile(path)
    out = []
    for sh in xl.sheet_names:
        n = norm(sh).lower()
        if not ("tesouro" in n or "renda fixa" in n or "renda_fixa" in n):
            continue
        df = pd.read_excel(path, sheet_name=sh)
        for _, r in df.iterrows():
            d = {c: r[c] for c in df.columns}
            cod = getcol(d, "Código", "Codigo", "Código do Ativo")
            prod = getcol(d, "Produto")
            qty = getcol(d, "Quantidade")
            inst = getcol(d, "Instituição", "Instituicao")
            if prod is None and cod is None:
                continue
            out.append({"codigo": (str(cod).strip() if cod is not None else ""),
                        "produto": (str(prod).strip() if prod is not None else ""),
                        "quantidade": (float(qty) if isinstance(qty, (int, float)) else None),
                        "corretora": (str(inst).strip() if inst is not None else "")})
    return out


def is_rf_title(key, descr=""):
    """An informe item that is a B3 renda-fixa TITLE (not a bank RDB/conta/fundo)."""
    k = norm(key).upper()
    if k.startswith("TESOURO"):
        return True
    if k.startswith(RF_PREFIX) and not k.startswith(("RDB", "NU-")):
        return True
    # security código pattern (e.g. CDBA254K3GV, 25J01903256, ENAT14) embedded
    return bool(re.search(r"\b[A-Z0-9]{6,}\b", k)) and any(p in (k + " " + norm(descr).upper()) for p in RF_PREFIX + ("TESOURO",))


def brl(x):
    try:
        return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return ""


def hdr(ws, n, color=PURPLE):
    for c in range(1, n + 1):
        cell = ws.cell(1, c)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=("FFFFFFFF" if color == PURPLE else "FF000000"))
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def sheet(wb, title, cols, rows, first=False):
    ws = wb.active if first else wb.create_sheet(title)
    ws.title = title
    ws.append(cols)
    hdr(ws, len(cols))
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, 1):
        w = max(len(str(c)), *(len(str(ws.cell(rr, i).value or "")) for rr in range(2, ws.max_row + 1))) if ws.max_row > 1 else len(str(c))
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 70)
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--posicao", required=True, help="B3 Posição export (for the RF position + validation)")
    ap.add_argument("--informes", required=True, help="processed/informes.json")
    ap.add_argument("--out", default="processed/fixed_income.xlsx")
    a = ap.parse_args()

    pos = read_rf_position(a.posicao)
    data = json.loads(Path(a.informes).read_text(encoding="utf-8"))
    # position security identity tokens (código + produto, uppercased) — to match informe items
    pos_tokens = set()
    for p in pos:
        if p["codigo"]:
            pos_tokens.add(norm(p["codigo"]).upper())
        if p["produto"]:
            pos_tokens.add(norm(p["produto"]).upper())

    def owns(it):
        key = norm(it.get("key")).upper()
        descr = norm(it.get("descr") or it.get("discriminacao") or "").upper()
        if key in pos_tokens:
            return True
        if any(tok and (tok in key or tok in descr or key in tok) for tok in pos_tokens):
            return True
        return is_rf_title(key, descr)

    # bens: RF titles among the informe bens (grupo 04). RDB/conta/fundo/exterior excluded.
    bens = [it for it in data.get("bens", [])
            if (it.get("grupo") == 4) and owns(it) and not norm(it.get("key")).upper().startswith(("NU-RDB", "NU-CONTA", "NU-RESERVA"))]
    # isentos código 12 = juros de CRA/CRI/LCI/LCA/debênture incentivada (sempre RF)
    isentos = [it for it in data.get("isentos", []) if int(it.get("codigo") or 0) == 12]
    # exclusiva código 06 que é título de RF (CDB/Tesouro) — RDB/come-cotas ficam no consolidate
    exclusiva = [it for it in data.get("exclusiva", []) if int(it.get("codigo") or 0) == 6 and owns(it)]

    owned_keys = {norm(it.get("key")).upper() for it in bens}

    # validation: every B3 renda-fixa position security must be covered by a bens item
    val_rows = []
    for p in pos:
        cod = norm(p["codigo"]).upper(); prod = norm(p["produto"]).upper()
        covered = any(k == cod or k == prod or cod in k or (k and k in prod) for k in owned_keys) \
                  or any(cod and cod in norm(b.get("descr") or "").upper() for b in bens)
        val_rows.append({**p, "validacao": "OK (no informe)" if covered
                         else "⚠️ FALTA no informe — confira o informe da corretora"})

    wb = Workbook()
    sheet(wb, "position", ["codigo", "produto", "quantidade", "corretora", "validacao"], val_rows, first=True)
    sheet(wb, "bens_e_direitos",
          ["key", "grupo", "codigo", "localizacao", "cnpj", "descr", "valor_2024", "valor_2025", "source"],
          [{**{"localizacao": 105}, **it} for it in bens])
    sheet(wb, "isentos", ["codigo", "key", "cnpj", "descr", "valor_2025", "source"], isentos)
    sheet(wb, "exclusiva", ["codigo", "key", "cnpj", "descr", "valor_2025", "source"], exclusiva)
    Path(a.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(a.out)

    tot_b = sum(float(b.get("valor_2025") or 0) for b in bens)
    miss = [v for v in val_rows if "FALTA" in v["validacao"]]
    print(f"OK -> {a.out}")
    print(f"  position   {len(pos)} títulos RF na B3 ({len(miss)} sem informe)")
    print(f"  bens_e_direitos  {len(bens)} itens, total R$ {brl(round(tot_b,2))}")
    print(f"  isentos    {len(isentos)} (cód 12)   exclusiva {len(exclusiva)} (cód 06)")
    if miss:
        print("  [!] titulos na Posicao B3 SEM item no informe (confira o informe):")
        for v in miss:
            print(f"     {v['codigo'] or v['produto']}  qtd {v['quantidade']}")


if __name__ == "__main__":
    main()
