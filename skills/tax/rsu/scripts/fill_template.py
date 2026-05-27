#!/usr/bin/env python3
"""Fill the RSU IRPF template (template_rsu.xlsx) from a data.json file.

Usage:
    python fill_template.py template_rsu.xlsx output.xlsx data.json

Writes ONLY the input cells and preserves every formula. The vesting input is
PER-GRANT (one row per grant per release date) in the "Vesting detail" table; the
4 by-date total rows (8-11) that the calc engine reads are SUMIFS/INDEX formulas, so
this script only sets their release dates. Dates are formatted yyyy-mm-dd and the
workbook is flagged to recalculate on open.

data.json shape (numbers are illustrative placeholders):
{
  "saldo_inicial": {"quantidade": 1000, "preco_medio_brl": 40.0, "custo_medio_usd": 8.0},
  "vesting": [   # one entry per grant per release date (the per-grant detail)
    {"release_date": "2025-01-03", "award_date": "2021-10-04", "award_number": "GRANT-A",
     "qty": 41, "price_usd": 10.0},
    ...
  ],
  "venda": [{"date": "2025-08-28", "qty": 50, "price_usd": 12.0}, ...]  # up to 10 rows (16-25)
}

Requires: openpyxl  (pip install openpyxl)
"""
import sys
import json
import datetime as dt

import openpyxl

INPUT_SHEET = "input"          # adjust if the template renames it
DATE_FMT = "yyyy-mm-dd"
BRL_FMT = "[$R$]#,##0.0000"
VEST_TOTAL_ROWS = (8, 11)      # the 4 by-date total rows the engine reads (release dates only)
DETAIL_START = 30              # first per-grant detail row
DETAIL_MAX = 45


def _date(iso):
    y, m, d = map(int, iso.split("-"))
    return dt.datetime(y, m, d)


def fill(template, output, data):
    wb = openpyxl.load_workbook(template, data_only=False)
    ws = wb[INPUT_SHEET]

    # 1) Opening balance (row 4)
    s = data.get("saldo_inicial")
    if s:
        ws["D4"] = s["quantidade"]
        ws["E4"] = s["preco_medio_brl"]
        ws["E4"].number_format = BRL_FMT       # template ships this cell as USD format
        ws["G4"] = s["custo_medio_usd"]
        ws["B4"].number_format = DATE_FMT

    # 2) Vesting — per-grant detail rows + by-date totals' release dates
    vest = data.get("vesting", [])
    if len(vest) > (DETAIL_MAX - DETAIL_START + 1):
        raise SystemExit(f"Too many vesting rows ({len(vest)}); detail holds "
                         f"{DETAIL_MAX - DETAIL_START + 1}. Extend the detail table.")
    for i, v in enumerate(vest):
        r = DETAIL_START + i
        ws[f"B{r}"] = _date(v["release_date"]); ws[f"B{r}"].number_format = DATE_FMT
        ws[f"C{r}"] = _date(v["award_date"]);   ws[f"C{r}"].number_format = DATE_FMT
        ws[f"D{r}"] = v["award_number"]
        ws[f"E{r}"] = v["qty"]
        ws[f"F{r}"] = v["price_usd"]
    # distinct release dates (chronological) -> the 4 by-date total rows the engine reads
    seen, dates = set(), []
    for v in vest:
        if v["release_date"] not in seen:
            seen.add(v["release_date"]); dates.append(v["release_date"])
    dates.sort()
    r0, r1 = VEST_TOTAL_ROWS
    if len(dates) > (r1 - r0 + 1):
        raise SystemExit(f"{len(dates)} distinct vesting dates but only {r1 - r0 + 1} "
                         f"total rows ({r0}-{r1}). The engine supports up to that many.")
    for i in range(r1 - r0 + 1):
        cell = f"B{r0 + i}"
        if i < len(dates):
            ws[cell] = _date(dates[i]); ws[cell].number_format = DATE_FMT
        else:
            ws[cell] = None            # clear unused total rows so they don't pull stale dates

    # 3) Sales (rows 16-25)
    for i, sale in enumerate(data.get("venda", [])[:10]):
        r = 16 + i
        ws[f"B{r}"] = _date(sale["date"]); ws[f"B{r}"].number_format = DATE_FMT
        ws[f"D{r}"] = sale["qty"]
        ws[f"E{r}"] = sale["price_usd"]

    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(output)


def main(argv):
    if len(argv) != 3:
        print(__doc__)
        return 1
    template, output, data_path = argv
    with open(data_path, encoding="utf-8") as f:
        data = json.load(f)
    fill(template, output, data)
    print(f"Filled {output} from {data_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
