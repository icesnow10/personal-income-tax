#!/usr/bin/env python3
"""/completeness — reconcile the B3 output against the broker/bank statements, BY IRPF FICHA.

ONE comparison, written to a single .md file, organized by the IRPF filling sections:
  - Bens e Direitos
  - Rendimentos Isentos e Não Tributáveis
  - Rendimentos Sujeitos à Tributação Exclusiva/Definitiva

Each section compares b3_source × informe_de_rendimentos line by line (with grupo/código and the
value of each side), flags divergences, and — crucially — gives a SUM reconciliation per código, so
a per-item "FALTA" that actually reconciles at the total (e.g. b3 folds IRDM11→IRIM11 while the
informe still lists them apart) is visible as a matching sum.

  - b3_source             = brazil_investments.xlsx (b3 skill output): sheets IRPF_bens_e_direitos,
                            IRPF_rendimentos_isentos, IRPF_rendimentos_exclusivos.
  - informe_de_rendimentos = the broker/bank statements, transcribed by the agent into informes.json.

Authority is per ficha. **Bens e Direitos:** b3_source wins on VALUE (custo = preço médio × qtd);
the informe wins on grupo/código (e.g. NUIF11 = 07/10 FI-Infra, not the 07/03 default). **Rendimentos
(isentos/exclusiva):** the INFORME wins on value AND classification — it is regime de caixa across
brokers + escriturador, while the b3 income sheets are partial and never declared. Mismatches flagged.

The informe layouts vary wildly, so EXTRACTION is the agent's job (read each PDF, normalize to
JSON); this script only matches, diffs, and writes the .md. Keep it generic.

Besides writing the report, `compare` also **audits and edits the deliverable** `irpf_consolidated.xlsx`
(the /generate output) in place: any value that drifts from its ficha authority (Bens → b3_source custo;
rendimentos → o informe) is rewritten and stamped in an `obs_completeness` column, and the same
adjustments are listed in the report (use --no-apply to audit without editing).

Subcommands
  extract <docs_dir>      Dump candidate Bens-e-Direitos lines from every PDF (a parsing aid).
  compare --investimentos brazil_investments.xlsx --informes informes.json
          [--out report.md] [--consolidado irpf_consolidated.xlsx] [--no-apply] [--tol 0.5]

informes.json — a list of objects, each tagged with its IRPF ficha:
  {"ficha":"bens|isentos|exclusiva", "key":"VALE3", "codigo":1, "grupo":3, "descr":"...",
   "cnpj":"...", "valor_2025":1374.20, "quantidade":20, "source":"btg"}
  - ficha "bens": grupo+codigo, valor_2025 (omit/null if the informe shows only quantidade, as for
    ações/FII), quantidade.
  - ficha "isentos"/"exclusiva": codigo (09/99 isentos; 10/06 exclusiva), key=ticker/fonte, valor_2025.

Requires: pandas, openpyxl, pypdf
"""
import argparse, glob, json, os, re, unicodedata
from pathlib import Path

VALUE_RE = re.compile(r"\d{1,3}(?:\.\d{3})*,\d{2}")
CNPJ_RE  = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-?\d{0,2}")
TICKER_RE = re.compile(r"\b[A-Z]{4}\d{1,2}\b")
RF_CODE_RE = re.compile(r"\b[A-Z]{2,4}\d[A-Z0-9]{4,}\b")


def norm(s):
    if s is None: return ""
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().strip()

def brl(x):
    try: return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError): return ""

def qfmt(q):
    if q is None: return ""
    return str(int(q)) if float(q) == int(q) else f"{q}"

# --- renda fixa: validação Posição B3 × informe (portada do antigo /fixed_income) ---
RF_PREFIX = ("CDB", "RDB", "CRA", "CRI", "CDCA", "DEB", "LCI", "LCA", "LF", "LIG", "LH", "LCD")

def _getcol(d, *names):
    for n in names:
        for k in d:
            if norm(k).lower() == norm(n).lower():
                return d[k]
    return None

def read_rf_position(path):
    """B3 Posição renda-fixa / Tesouro sheets → [{codigo, produto, quantidade, corretora}]."""
    if not path or not os.path.exists(path):
        return []
    import pandas as pd
    xl = pd.ExcelFile(path)
    out = []
    for sh in xl.sheet_names:
        n = norm(sh).lower()
        if not ("tesouro" in n or "renda fixa" in n or "renda_fixa" in n):
            continue
        df = pd.read_excel(path, sheet_name=sh)
        def _txt(v):                                       # NaN/None → "" (planilhas trazem linhas em branco)
            return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()
        for _, r in df.iterrows():
            d = {c: r[c] for c in df.columns}
            cod = _txt(_getcol(d, "Código", "Codigo", "Código do Ativo"))
            prod = _txt(_getcol(d, "Produto"))
            qty = _getcol(d, "Quantidade")
            inst = _txt(_getcol(d, "Instituição", "Instituicao"))
            if not prod and not cod:                       # linha em branco / subtotal → pula
                continue
            qok = isinstance(qty, (int, float)) and not pd.isna(qty)
            out.append({"codigo": cod, "produto": prod,
                        "quantidade": (float(qty) if qok else None),
                        "corretora": inst})
    return out

def rf_position_check(posicao_path, informes_path):
    """Every renda-fixa security in the B3 Posição must have a matching bens item in informes.json
    (the value is declared from the informe). A missing one = incomplete /read transcription."""
    pos = read_rf_position(posicao_path)
    if not pos:
        return []
    data = json.loads(Path(informes_path).read_text(encoding="utf-8"))
    bens = [it for it in (data.get("bens") or [])
            if (it.get("grupo") == 4) and not norm(it.get("key")).upper().startswith(("NU-RDB", "NU-CONTA", "NU-RESERVA"))]
    owned = {norm(it.get("key")).upper(): norm(it.get("descr") or it.get("discriminacao") or "").upper() for it in bens}
    rows = []
    for p in pos:
        cod = norm(p["codigo"]).upper(); prod = norm(p["produto"]).upper()
        covered = any(k == cod or k == prod or (cod and cod in k) or (k and k in prod) for k in owned) \
                  or any(cod and cod in dsc for dsc in owned.values())
        rows.append({**p, "covered": covered})
    return rows

def foreign_keys(informes_path):
    """Keys of foreign Bens e Direitos (b3:false, localizacao != 105) — to flag foreign income that
    landed in isentos/exclusiva (it belongs in OTHER fichas: carnê-leão / GCAP)."""
    data = json.loads(Path(informes_path).read_text(encoding="utf-8"))
    out = set()
    for it in (data.get("bens") or []):
        loc = norm(it.get("localizacao"))
        if not it.get("b3") and loc not in ("", "105", "0", "None"):
            out.add(norm(it.get("key")).upper())
    return out

def to_num(s):
    if s is None: return None
    if isinstance(s, (int, float)):
        f = float(s)
        return None if f != f else f                # NaN (pandas empty cell) -> None
    if str(s).strip().lower() in ("nan", "none"): return None
    t = str(s).strip().replace("R$", "").strip()
    if not t: return None
    t = t.replace(".", "").replace(",", ".") if "," in t else t
    try: return float(t)
    except ValueError: return None

def eq(a, b, tol):
    if a is None or b is None: return a is None and b is None
    return abs(a - b) <= max(tol, abs(a) * 0.005)

def gint(v):
    try: return int(float(v))
    except (TypeError, ValueError): return None


# --------------------------- extract (parsing aid) ---------------------------
def pdf_text(path):
    from pypdf import PdfReader
    try:
        r = PdfReader(path)
        return "\n".join((p.extract_text() or "") for p in r.pages)
    except Exception as e:
        return f"<<ERRO ao ler PDF: {e}>>"

def extract(docs_dir):
    pdfs = sorted(glob.glob(os.path.join(docs_dir, "*.pdf"))) + sorted(glob.glob(os.path.join(docs_dir, "*.PDF")))
    if not pdfs:
        print(f"Nenhum PDF em {docs_dir}"); return
    for f in pdfs:
        txt = pdf_text(f)
        print(f"\n################ {os.path.basename(f)}")
        if txt.startswith("<<ERRO"):
            print("  ", txt); continue
        emitted = 0
        for raw in txt.splitlines():
            s = raw.strip()
            if not s: continue
            if VALUE_RE.search(s) and (CNPJ_RE.search(s) or TICKER_RE.search(s)
                    or RF_CODE_RE.search(s) or re.search(
                        r"(saldo|cota|fundo|tesouro|a[cç][aã]o|reserva|conta|moeda|"
                        r"CDB|CRA|CRI|DEB|LC[IA]|RDB|BDR|ETF)", s, re.I)):
                print("  ", s[:140]); emitted += 1
        if not emitted:
            print("  (nenhuma linha candidata — imagem/cifrado; transcreva manualmente)")


# --------------------------- load b3_source (3 IRPF sheets) ---------------------------
def load_b3(path):
    import pandas as pd
    xl = pd.ExcelFile(path)
    def sheet(part):
        nm = next((s for s in xl.sheet_names if part in s.lower()), None)
        return pd.read_excel(path, sheet_name=nm) if nm else None
    def year_col(df):
        ys = [c for c in df.columns if str(c).isdigit() and len(str(c)) == 4]
        if ys: return max(ys, key=lambda c: int(str(c)))
        # 'valor_YYYY' columns: pick the LATEST year (valor_2025), not the first one seen (era valor_2024)
        vy = [(int(str(c)[-4:]), c) for c in df.columns
              if str(c).lower().startswith("valor_") and str(c)[-4:].isdigit()]
        if vy: return max(vy)[1]
        cols = {str(c).lower(): c for c in df.columns}
        return cols.get("valor")

    out = {"bens": {}, "isentos": {}, "exclusiva": {}}
    bd = sheet("bens_e_direitos")
    if bd is not None:
        cols = {str(c).lower(): c for c in bd.columns}
        vcol = year_col(bd)
        for _, r in bd.iterrows():
            tk = norm(r[cols.get("ticker")]).upper()
            if not tk: continue
            descr = str(r.get(cols.get("discriminacao", ""), "") or "")
            mq = re.search(r"([\d.]+)\s*UNIDADES", descr, re.I)
            out["bens"][tk] = {"grupo": gint(r.get(cols.get("grupo"))), "codigo": gint(r.get(cols.get("codigo"))),
                               "valor": to_num(r.get(vcol)), "qtd": to_num(mq.group(1)) if mq else None,
                               "descr": descr}
    for part, key in (("rendimentos_isentos", "isentos"), ("rendimentos_exclusiv", "exclusiva")):
        df = sheet(part)
        if df is None: continue
        cols = {str(c).lower(): c for c in df.columns}
        vcol = year_col(df)
        tcol = cols.get("tipo_de_rendimento"); kcol = cols.get("ticker")
        for _, r in df.iterrows():
            tk = norm(r[kcol]).upper(); cod = gint(r[tcol]); v = to_num(r[vcol])
            if tk and cod is not None and v is not None:
                out[key][(cod, tk)] = {"codigo": cod, "valor": v, "cnpj": norm(r.get(cols.get("cnpj"), ""))}
    # ticker renames the b3 skill applied (incorporações, BDR renames) — read from aux_mapping so
    # the informe can be folded the same way (e.g. IRDM11→IRIM11).
    ren = {}
    am = sheet("aux_mapping")
    if am is not None:
        cols = {str(c).lower(): c for c in am.columns}
        fc, tc = cols.get("from_ticker"), cols.get("to_ticker")
        if fc is not None and tc is not None:
            for _, r in am.iterrows():
                f, t = norm(r.get(fc)).upper(), norm(r.get(tc)).upper()
                if f and t: ren[f] = t
    out["_renames"] = ren
    return out


def load_informes(path):
    """Read the unified informes.json that /read produces: a single object with bens/isentos/exclusiva
    lists (the SAME file /generate consumes). For each item: key, grupo/codigo, valor_2025, quantidade,
    source (the PDF the value came from)."""
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    out = {"bens": {}, "isentos": {}, "exclusiva": {}}
    def mkrec(it):
        key = norm(it.get("key")).upper()
        rec = {"key": key, "grupo": gint(it.get("grupo")), "codigo": gint(it.get("codigo")),
               "descr": it.get("descr", ""), "cnpj": norm(it.get("cnpj")),
               "valor": to_num(it.get("valor_2025")), "quantidade": to_num(it.get("quantidade")),
               "source": it.get("source", "")}
        rec["componentes"] = [(key, rec["valor"])] if rec["valor"] is not None else []
        return key, rec
    for it in data.get("bens", []):
        key, rec = mkrec(it)
        out["bens"][key] = rec
    for ficha in ("isentos", "exclusiva"):
        for it in data.get(ficha, []):
            key, rec = mkrec(it)
            out[ficha][(rec["codigo"], key)] = rec
    return out


def parse_md_table(path):
    """Tiny markdown-table reader: returns list of dicts keyed by the (lowercased) header cells.
    Ignores lines outside the first pipe-table and the |---| separator."""
    rows, header = [], None
    for ln in Path(path).read_text(encoding="utf-8").splitlines():
        s = ln.strip()
        if not (s.startswith("|") and s.endswith("|")): continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if set("".join(cells)) <= set("-: "):            # separator row
            continue
        if header is None:
            header = [c.lower() for c in cells]; continue
        if len(cells) == len(header):
            rows.append(dict(zip(header, cells)))
    return rows


def load_escriturador(path):
    """ticker(UPPER) -> {escriturador, cnpj, tag, source}. The `tag` is the lowercase substring expected
    in a rendimento `source` when THAT escriturador's informe was actually used (e.g. 'bradesco')."""
    if not path or not os.path.exists(path): return {}
    out = {}
    for r in parse_md_table(path):
        tk = norm(r.get("ticker")).upper()
        if not tk or tk.startswith("EXEMPLO"): continue
        tag = (r.get("tag") or r.get("escriturador") or "").strip().lower()
        out[tk] = {"escriturador": r.get("escriturador", "").strip(),
                   "cnpj": r.get("cnpj", "").strip(), "tag": tag, "source": r.get("source", "").strip()}
    return out


def _merge(a, b):
    """Fold informe item b into a (same b3 ticker after a rename) — sum valor/quantidade, keep the
    component breakdown and union the PDF sources, for traceability of the aglutination."""
    for f in ("valor", "quantidade"):
        if b.get(f) is not None: a[f] = (a.get(f) or 0.0) + b[f]
    parts = []
    for s in (a.get("source"), b.get("source")):
        for p in (s or "").split(" + "):
            if p and p not in parts: parts.append(p)
    a["source"] = " + ".join(parts)
    a.setdefault("componentes", []).extend(b.get("componentes", []))

def apply_renames(inf, renames):
    """Fold informe keys the SAME way the b3 skill does (ticker_memory renames, read from the
    workbook's aux_mapping): e.g. an incorporation IRDM11→IRIM11 aglutinates both informe lines into
    one IRIM11 so it reconciles with b3_source's folded position."""
    rn = lambda k: renames.get(k, k)
    out = {"bens": {}, "isentos": {}, "exclusiva": {}}
    for k, rec in inf["bens"].items():
        nk = rn(k); rec = dict(rec)
        if nk in out["bens"]: _merge(out["bens"][nk], rec)
        else: rec["key"] = nk; out["bens"][nk] = rec
    for ficha in ("isentos", "exclusiva"):
        for (cod, k), rec in inf[ficha].items():
            nk = rn(k); rec = dict(rec); key = (cod, nk)
            if key in out[ficha]: _merge(out[ficha][key], rec)
            else: out[ficha][key] = rec
    return out


# --------------------------- audit + fix the consolidated workbook ---------------------------
def audit_consolidado(path, b3, inf, tol, apply=True):
    """Audit irpf_consolidated.xlsx (the /generate deliverable) against the per-ficha authority and,
    when apply=True, EDIT it in place so every value matches the authority — Bens e Direitos → b3_source
    (custo); rendimentos → o informe. Every row that diverged from its other source is stamped in an
    `obs_completeness` column (both the value actually used and the value the other source carried), so
    the adjustment is visible in the deliverable too. Returns the list of (sheet, key, kind, msg)."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    sheets = {"bens_e_direitos": "bens", "isentos": "isentos", "exclusiva_definitiva": "exclusiva"}
    adj = []
    for sname, ficha in sheets.items():
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        header = [(c.value if c.value is not None else "") for c in ws[1]]
        hmap = {str(h).strip().lower(): i for i, h in enumerate(header)}
        vcol = hmap.get("valor_2025")
        if vcol is None: continue
        ocol = hmap.get("obs_completeness")
        if ocol is None:
            ocol = len(header); ws.cell(1, ocol + 1, "obs_completeness")
        for ri in range(2, ws.max_row + 1):
            if ws.cell(ri, 1).value == "TOTAL": continue
            cur = to_num(ws.cell(ri, vcol + 1).value)
            if ficha == "bens":
                if str(ws.cell(ri, (hmap.get("origem", 0)) + 1).value or "") != "B3":
                    continue                                   # não-B3: só existe no informe, nada a conferir
                disc = str(ws.cell(ri, (hmap.get("discriminacao", 0)) + 1).value or "").upper()
                k = next((t for t in b3["bens"] if t and t in disc), None)
                if k is None: continue
                auth = (b3["bens"].get(k) or {}).get("valor")   # custo b3_source
                other = (inf["bens"].get(k) or {}).get("valor") # valor do informe (cross-check)
                who, other_name = "b3_source", "informe"
            else:
                cod = gint(ws.cell(ri, (hmap.get("codigo")) + 1).value)
                key = norm(ws.cell(ri, (hmap.get("key")) + 1).value).upper()
                auth = (inf[ficha].get((cod, key)) or {}).get("valor")   # informe = autoridade
                other = (b3[ficha].get((cod, key)) or {}).get("valor")   # aba income (parcial) do b3
                who, other_name = "informe", "aba income b3"
            if auth is None: continue
            note = ""
            if cur is not None and not eq(cur, auth, tol) and apply:
                ws.cell(ri, vcol + 1, round(auth, 2))
                note = f"AJUSTADO por /completeness: R$ {brl(cur)} -> R$ {brl(auth)} (vale {who})"
                adj.append((sname, k if ficha == "bens" else key, "ajuste", note))
            elif cur is not None and not eq(cur, auth, tol) and not apply:
                note = f"DIVERGE do consolidado (R$ {brl(cur)}); deveria ser R$ {brl(auth)} (vale {who})"
                adj.append((sname, k if ficha == "bens" else key, "diverge", note))
            elif other is not None and not eq(auth, other, tol):
                note = f"conferido /completeness: vale {who} R$ {brl(auth)} ({other_name} trazia R$ {brl(other)})"
                adj.append((sname, k if ficha == "bens" else key, "conferido", note))
            if note: ws.cell(ri, ocol + 1, note)
        # recompute TOTAL (valor_2025) if values were edited
        last = ws.max_row
        if ws.cell(last, 1).value == "TOTAL":
            s = sum(v for ri in range(2, last)
                    if (v := to_num(ws.cell(ri, vcol + 1).value)) is not None)
            ws.cell(last, vcol + 1, round(s, 2))
    if apply or any(a[2] == "diverge" for a in adj):
        wb.save(path)
    return adj


# --------------------------- the report ---------------------------
def looks_b3(key):
    return bool(TICKER_RE.fullmatch(key) or RF_CODE_RE.fullmatch(key)
                or key.startswith(("CDB", "CRA", "CRI", "DEB", "LCI", "LCA", "RDB")) or key.startswith("TESOURO"))

def status_val(vb, vi, tol, authority="b3_source"):
    """authority = quem vence a divergência de VALOR nesta ficha. Bens e Direitos → b3_source
    (custo = preço médio × qtd). Rendimentos (isentos/exclusiva) → o informe (regime de caixa,
    soma corretoras + escriturador; a aba de income do b3 é parcial e não é usada)."""
    if vb is None and vi is None: return "—"
    if vb is None: return None  # handled by caller (informe-only)
    if vi is None: return None  # handled by caller (b3-only / informe sem valor)
    if eq(vb, vi, tol): return "OK"
    vale = "vale o informe" if authority == "informe" else "vale b3_source"
    return f"DIVERGE (b3 {brl(vb)} × informe {brl(vi)} — {vale})"


def compare(args):
    b3 = load_b3(args.investimentos)
    inf = apply_renames(load_informes(args.informes), b3.get("_renames", {}))
    esc_path = args.escriturador_memory or os.path.join("memory", "escriturador_memory.md")
    esc = load_escriturador(esc_path)
    md = args.out if args.out.lower().endswith(".md") else os.path.splitext(args.out)[0] + ".md"

    L = ["# Completeness — b3_source × informe_de_rendimentos (por ficha do IRPF)\n",
         "**b3_source** = `brazil_investments.xlsx` · **informe_de_rendimentos** = informes das corretoras/bancos. "
         "Autoridade por ficha: em **Bens e Direitos** o **valor** vem do b3_source (custo = preço médio × qtd) "
         "e o **grupo/código** vem do informe; em **Rendimentos Isentos** e **Tributação Exclusiva** o **valor "
         "e a classificação** vêm do **informe** (regime de caixa, somando corretoras + escriturador — a aba de "
         "income do b3 é parcial e não é declarada). Não é orientação fiscal.\n"]
    counts = {}
    def bump(k): counts[k] = counts.get(k, 0) + 1

    # ---------------- 1. Bens e Direitos ----------------
    L += ["## Bens e Direitos\n",
          "| grupo | código | asset | qtd b3 | qtd informe | valor b3 | valor informe | fonte (PDF) | status |",
          "|---|---|---|--:|--:|--:|--:|---|---|"]
    keys = sorted(set(b3["bens"]) | set(inf["bens"]),
                  key=lambda k: (k not in b3["bens"], k))   # B3 (in b3_source) first
    total_decl = 0.0; n_naob3 = 0
    for k in keys:
        a = b3["bens"].get(k); b = inf["bens"].get(k)
        grupo = (b or {}).get("grupo") or (a or {}).get("grupo") or ""
        cb, ci = (a or {}).get("codigo"), (b or {}).get("codigo")
        codigo = ci or cb or ""
        cod_div = f" ⚠️b3={cb}/inf={ci}" if (a and b and cb is not None and ci is not None and cb != ci) else ""
        vb, vi = (a or {}).get("valor"), (b or {}).get("valor")
        qb, qi = (a or {}).get("qtd"), (b or {}).get("quantidade")
        decl = vb if vb is not None else vi            # valor a declarar: b3 p/ B3, informe p/ não-B3
        if decl is not None: total_decl += decl
        if b and not a: n_naob3 += 1
        qdiv = (qb is not None and qi is not None and abs(qb - qi) > 1e-6)
        if a and not b:
            st = "só no b3_source (não no informe)"
        elif b and not a:
            st = "FALTA em b3_source (B3!)" if looks_b3(k) else "só no informe (não-B3 — declarar à mão)"
        elif vi is None and qi is not None:
            st = (f"qtd DIVERGE (b3 {qfmt(qb)} × informe {qfmt(qi)})" if qdiv
                  else f"OK (informe só qtd {qfmt(qi)}; valor = custo b3_source)")
        else:
            st = status_val(vb, vi, args.tol) or "—"
            if qdiv: st += f" | qtd DIVERGE (b3 {qfmt(qb)} × inf {qfmt(qi)})"
        bump(st.split(" (")[0].split(" —")[0].split(" |")[0])
        fonte = (b or {}).get("source", "")
        L.append(f"| {grupo} | {codigo}{cod_div} | {k} | {qfmt(qb)} | {qfmt(qi)} | {brl(vb)} | {brl(vi)} | {fonte} | {st} |")
    L.append(f"| | | **TOTAL a declarar** (b3_source p/ B3 + informe p/ não-B3) | | | | **{brl(total_decl)}** | | "
             f"**{n_naob3} itens não-B3 a incluir à mão** |")
    L.append("")

    # ---------------- 2 & 3. Rendimentos (isentos / exclusiva) ----------------
    recl_b3, recl_inf = [], []                              # itens "só num lado" p/ detectar reclassificação cruzada
    sec_titles = {"isentos": "Rendimentos Isentos e Não Tributáveis",
                  "exclusiva": "Rendimentos Sujeitos à Tributação Exclusiva/Definitiva"}
    for ficha in ("isentos", "exclusiva"):
        L += [f"## {sec_titles[ficha]}\n"]
        bd, ind = b3[ficha], inf[ficha]
        codigos = sorted({c for (c, _) in bd} | {c for (c, _) in ind})
        for cod in codigos:
            L += [f"### Código {cod:02d}\n",
                  "| código | asset | b3_source | informe | fonte (PDF) | status |", "|---|---|--:|--:|---|---|"]
            ks = sorted({tk for (c, tk) in bd if c == cod} | {tk for (c, tk) in ind if c == cod})
            s_b3 = s_inf = 0.0
            for tk in ks:
                vb = (bd.get((cod, tk)) or {}).get("valor")
                ri = ind.get((cod, tk)) or {}
                vi = ri.get("valor")
                fonte = ri.get("source", "")
                comps = ri.get("componentes", [])
                aglut = (" [aglutinado: " + " + ".join(f"{ck} {brl(cv)}" for ck, cv in comps) + "]") if len(comps) > 1 else ""
                if vb is not None: s_b3 += vb
                if vi is not None: s_inf += vi
                if vb is not None and vi is None:
                    st = "só no b3_source"; recl_b3.append((ficha, cod, tk, vb))
                elif vi is not None and vb is None:
                    st = "só no informe"; recl_inf.append((ficha, cod, tk, vi))
                else: st = status_val(vb, vi, args.tol, authority="informe") or "—"
                bump(st.split(" (")[0])
                L.append(f"| {cod:02d} | {tk}{aglut} | {brl(vb)} | {brl(vi)} | {fonte} | {st} |")
            sd = ("OK (somatório bate)" if eq(s_b3, s_inf, args.tol)
                  else f"DIVERGE (dif {brl(s_b3 - s_inf)}) — declarar o total do informe R$ {brl(s_inf)}")
            L.append(f"| {cod:02d} | **SOMA código {cod:02d}** | **{brl(s_b3)}** | **{brl(s_inf)}** | | **{sd}** |")
            L.append("")

    # ---------------- escriturador das ações/FII (o informe do escriturador foi usado?) ----------------
    # O escriturador emite o informe AUTORITATIVO de dividendos/JCP do papel (mais que a corretora).
    # Confere, por ação/FII, se o `source` do rendimento referencia o escriturador (memory tag).
    equities = {tk: r for tk, r in b3["bens"].items()
                if r.get("grupo") in (3, 7) or (r.get("grupo") == 4 and r.get("codigo") == 4)}
    def tipo_b3(g, c):
        if g == 3: return "ação (B3 empresas-listadas)"
        if g == 7 and c == 10: return "FI-Infra (B3 fi-infra-listados)"
        if g == 7: return "FII (B3 fiis-listados)"
        if g == 4 and c == 4: return "BDR (B3 bdrs-patrocinados / nao-patrocinados-listados)"
        return "outro tipo (buscar o escriturador na B3 ou por busca web)"
    esc_acts = []
    if equities:
        L += ["## Escriturador das ações/FII (o informe do escriturador foi usado?)\n",
              "_O escriturador é a autoridade dos dividendos/JCP do papel. Conferência: alguma fonte do "
              "rendimento é o informe do escriturador? Preencha `memory/escriturador_memory.md` a partir da "
              "B3 ([ações](https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/empresas-listadas.htm) · "
              "[FIIs](https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/fundos-de-investimentos/fii/fiis-listados/) · "
              "[FI-Infra](https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/fundos-de-investimentos/fi-infra/fi-infra-listados/) · "
              "[BDR patrocinado](https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/bdrs/bdrs-patrocinados/bdrs-patrocinados-listados/) · "
              "[BDR não patrocinado](https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/bdrs/bdrs-nao-patrocinados/bdrs-nao-patrocinados-listados/))._\n",
              "| ticker | grupo | escriturador | usou o informe do escriturador? | rendimentos (fontes) |",
              "|---|--:|---|---|---|"]
        for tk in sorted(equities):
            grupo = equities[tk].get("grupo")
            srcs = []
            for ficha in ("isentos", "exclusiva"):
                for (c, k), rec in inf[ficha].items():
                    if k == tk and rec.get("source"): srcs.append(rec["source"])
            srcs_str = " · ".join(dict.fromkeys(srcs)); src_low = srcs_str.lower()
            e = esc.get(tk)
            if not e:
                tipo = tipo_b3(equities[tk].get("grupo"), equities[tk].get("codigo"))
                nome, st = "—", f"⚠️ escriturador desconhecido — PROCURAR ({tipo}) e preencher escriturador_memory.md"
                bump("escriturador desconhecido")
                esc_acts.append(f"- [ ] **`{tk}`**: escriturador desconhecido — **procurar o escriturador** "
                                f"em {tipo}, preencher `memory/escriturador_memory.md` e re-rodar.")
            else:
                nome = e["escriturador"] or "—"
                if not srcs:
                    st = "— (sem rendimentos no ano — nada a conferir)"
                elif e["tag"] and e["tag"] in src_low:
                    st = f"✅ sim ({e['tag']})"; bump("escriturador conferido")
                else:
                    st = (f"⚠️ NÃO — rendimentos vieram só de corretora; buscar o informe do escriturador "
                          f"({nome})"); bump("escriturador NÃO conferido")
                    esc_acts.append(f"- [ ] **`{tk}`**: rendimento sem o informe do escriturador ({nome}) — "
                                    f"buscar esse informe (dividendos/JCP podem estar incompletos, ex.: BBSE3).")
            L.append(f"| {tk} | {grupo} | {nome} | {st} | {srcs_str} |")
        L.append("")

    # ---------------- divergências de classificação (mesmo ativo+valor, código diferente) ----------------
    # ex.: BBSE3 7,41 — b3 mapeia "Rendimento" de ação para isento/99, mas o informe (Nubank) classifica
    # como JCP/exclusiva/10. Mesmo valor em código diferente = reclassificação, não valor faltando.
    recl = []
    for (fb, cb, tkb, vb) in recl_b3:
        for (fi, ci, tki, vi) in recl_inf:
            if tkb == tki and (cb != ci or fb != fi) and eq(vb, vi, args.tol):
                recl.append((tkb, vb, fb, cb, fi, ci))
    if recl:
        L += ["## Divergências de classificação (mesmo ativo e valor, código diferente)\n",
              "_Mesmo valor aparece em código diferente entre b3_source e informe — reclassificação, "
              "não valor faltando. O informe é a autoridade de classificação._\n",
              "| ativo | valor | b3_source | informe | ação |", "|---|--:|---|---|---|"]
        for (tk, v, fb, cb, fi, ci) in recl:
            L.append(f"| {tk} | {brl(v)} | {fb} / cód {cb:02d} | {fi} / cód {ci:02d} | usar o código do informe |")
            bump("reclassificação (código b3 × informe)")
        L.append("")

    # ---------------- renda fixa: Posição B3 × informe (validação migrada do /fixed_income) ----------------
    extra_acts = []
    if getattr(args, "posicao", None):
        rfp = rf_position_check(args.posicao, args.informes)
        if rfp:
            L += ["## Renda fixa — Posição B3 × informe\n",
                  "_Todo título de renda fixa custodiado na B3 deve ter um item correspondente no "
                  "`informes.json` (o valor de Bens e Direitos vem do informe). Faltando = transcrição "
                  "incompleta no /read — confira o informe da corretora e volte ao /read._\n",
                  "| código | produto | qtd | corretora | status |", "|---|---|--:|---|---|"]
            for r in rfp:
                ok = r["covered"]
                st = "OK (no informe)" if ok else "⚠️ FALTA no informe — confira o informe da corretora"
                L.append(f"| {r['codigo']} | {r['produto']} | {qfmt(r['quantidade'])} | {r['corretora']} | {st} |")
                bump("renda fixa coberta pelo informe" if ok else "renda fixa FALTA no informe")
                if not ok:
                    extra_acts.append(f"- [ ] **`{r['codigo'] or r['produto']}`** (renda fixa, qtd {qfmt(r['quantidade'])}): "
                                      f"na Posição B3 mas sem item no `informes.json` — buscar o informe e transcrever (/read).")
            L.append("")

    # ---------------- exterior: rendimento estrangeiro caiu em isento/exclusiva? ----------------
    fk = foreign_keys(args.informes)
    if fk:
        flagged = [(ficha, c, tk, (rec or {}).get("valor")) for ficha in ("isentos", "exclusiva")
                   for (c, tk), rec in inf[ficha].items() if norm(tk).upper() in fk]
        if flagged:
            L += ["## Exterior — rendimento em ficha errada?\n",
                  "_Estes rendimentos têm `key` de um bem no EXTERIOR (localização ≠ 105). Dividendo/ganho "
                  "no exterior é **tributável** (carnê-leão / GCAP / Tributáveis Recebidos do Exterior), "
                  "**não** entra em isento/exclusiva. Confira a ficha._\n",
                  "| ficha (atual) | código | ativo | valor | ação |", "|---|---|---|--:|---|"]
            for (ficha, c, tk, v) in flagged:
                L.append(f"| {ficha} | {c} | {tk} | {brl(v)} | mover p/ ficha de exterior (tributável) |")
                bump("rendimento de exterior em ficha errada")
                extra_acts.append(f"- [ ] **`{tk}`** (rendimento {brl(v)}): é do exterior — declarar em carnê-leão/GCAP/"
                                  f"Tributáveis do Exterior, não em {ficha}.")
            L.append("")

    # ---------------- auditar + editar o irpf_consolidated.xlsx ----------------
    cons = args.consolidado or os.path.join(os.path.dirname(os.path.abspath(md)) or ".", "irpf_consolidated.xlsx")
    adj = []
    if os.path.exists(cons):
        adj = audit_consolidado(cons, b3, inf, args.tol, apply=not args.no_apply)
        edits = [a for a in adj if a[2] == "ajuste"]
        divs  = [a for a in adj if a[2] == "diverge"]
        confs = [a for a in adj if a[2] == "conferido"]
        L += [f"## Ajustes no `{os.path.basename(cons)}`\n",
              ("Auditoria por ficha **editou** o consolidado para alinhar à autoridade " if not args.no_apply
               else "Auditoria por ficha (modo `--no-apply`, **sem editar**) ") +
              "(Bens e Direitos → b3_source; rendimentos → informe). "
              "Cada linha conferida/ajustada leva uma nota na coluna `obs_completeness` do próprio arquivo.\n"]
        if edits:
            L += [f"**{len(edits)} valor(es) ajustado(s):**", "", "| aba | item | ajuste |", "|---|---|---|"]
            L += [f"| {s} | {k} | {m.split(': ',1)[-1]} |" for (s, k, _, m) in edits]; L.append("")
            for a in edits: bump("ajuste no consolidado")
        if divs:
            L += [f"**{len(divs)} divergência(s) NÃO aplicada(s)** (rode sem `--no-apply` para corrigir):",
                  "", "| aba | item | divergência |", "|---|---|---|"]
            L += [f"| {s} | {k} | {m} |" for (s, k, _, m) in divs]; L.append("")
            for a in divs: bump("diverge no consolidado (não aplicado)")
        if confs:
            L += [f"**{len(confs)} valor(es) conferido(s)** (já corretos; divergiam da outra fonte):",
                  "", "| aba | item | conferência |", "|---|---|---|"]
            L += [f"| {s} | {k} | {m.split(': ',1)[-1]} |" for (s, k, _, m) in confs]; L.append("")
        if not adj:
            L += ["Nenhuma divergência: o consolidado já reflete a autoridade de cada ficha.\n"]

    # ---------------- resumo + action items ----------------
    L += ["## Resumo\n"] + [f"- **{v}** {k}" for k, v in sorted(counts.items(), key=lambda x: -x[1])]
    L += ["", "## Action items\n"]
    acts = []
    for k in keys:
        a, b = b3["bens"].get(k), inf["bens"].get(k)
        if a and b and a.get("codigo") is not None and b.get("codigo") is not None and a["codigo"] != b["codigo"]:
            acts.append(f"- [ ] **`{k}`** (Bens e Direitos): código diverge — b3_source={a['codigo']} × informe={b['codigo']} (vale o informe).")
        if b and not a and not looks_b3(k):
            acts.append(f"- [ ] **`{k}`** (Bens e Direitos): não-B3 — declarar à mão a partir do informe ({brl(b.get('valor'))}).")
        if b and not a and looks_b3(k):
            acts.append(f"- [ ] **`{k}`** (Bens e Direitos): ativo B3 ausente do b3_source — checar /b3.")
    for (tk, v, fb, cb, fi, ci) in recl:
        acts.append(f"- [ ] **`{tk}`** (rendimento {brl(v)}): reclassificação — b3 {fb}/cód {cb:02d} × "
                    f"informe {fi}/cód {ci:02d}; declarar no código do informe.")
    acts += esc_acts + extra_acts
    if not acts: acts = ["- [x] Sem pendências estruturais (confira as divergências de valor nas tabelas)."]
    L += acts

    Path(md).write_text("\n".join(L) + "\n", encoding="utf-8")
    # console digest
    print("=== Resumo (status) ===")
    for k, v in sorted(counts.items(), key=lambda x: -x[1]): print(f"  {v:3} {k}")
    if os.path.exists(cons):
        n_e = sum(1 for a in adj if a[2] == "ajuste"); n_d = sum(1 for a in adj if a[2] == "diverge")
        n_c = sum(1 for a in adj if a[2] == "conferido")
        verbo = "ajustes aplicados" if not args.no_apply else "divergências (não aplicadas)"
        print(f"\n{os.path.basename(cons)}: {n_e} {verbo}, {n_c} conferidos" + (f", {n_d} divergentes" if n_d else ""))
    print(f"\nrelatório -> {md}")


def main():
    ap = argparse.ArgumentParser(description="Completeness por ficha do IRPF: b3_source × informe_de_rendimentos.")
    sub = ap.add_subparsers(dest="cmd", required=True)
    pe = sub.add_parser("extract"); pe.add_argument("docs_dir")
    pc = sub.add_parser("compare")
    pc.add_argument("--investimentos", required=True, help="b3_source: brazil_investments.xlsx")
    pc.add_argument("--informes", required=True, help="informe_de_rendimentos: informes.json (com 'ficha' por item)")
    pc.add_argument("--posicao", help="B3 Posição export — valida que todo título de RF na B3 tem item no informe")
    pc.add_argument("--out", default="completeness_report.md")
    pc.add_argument("--consolidado", help="irpf_consolidated.xlsx a auditar/editar (default: ao lado do --out)")
    pc.add_argument("--escriturador-memory", help="memory/escriturador_memory.md (ticker -> escriturador; default: memory/escriturador_memory.md)")
    pc.add_argument("--no-apply", action="store_true", help="só auditar (não editar o consolidado); divergências são listadas")
    pc.add_argument("--tol", type=float, default=0.5)
    a = ap.parse_args()
    if a.cmd == "extract": extract(a.docs_dir)
    else: compare(a)


if __name__ == "__main__":
    main()
