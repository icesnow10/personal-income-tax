#!/usr/bin/env python3
"""fetch_escriturador — popula memory/escriturador_memory.md direto da B3 (sem prints, sem informes).

O escriturador (instituição depositária das cotas/ações) é a AUTORIDADE de dividendos/JCP do papel.
Em vez de procurar à mão, este script consulta a API pública da B3 para CADA ativo do
brazil_investments.xlsx (a saída do skill b3, que sempre existe) e escreve o escriturador:

  - Ações / BDR (grupo 3 / 4-4): listedCompaniesProxy/CompanyCall/GetListedSupplementCompany
        payload {"issuingCompany":"<código sem dígitos>","language":"pt-br"}  -> campo `hasCommom`
  - FII / FI-Infra (grupo 7): fundsProxy/fundsCall/GetListedSupplementFunds
        payload {"cnpj":"<8 díg.>","identifierFund":"<código sem dígitos>","typeFund":7} -> campo `ifd`

É o MESMO dado que aparece em B3 > "Informações Gerais do Fundo" > Escriturador (campo `ifd`/`hasCommom`).
Escreve `ticker | escriturador | cnpj | tag | source` em memory/escriturador_memory.md; o /completeness
lê esse arquivo. Preserva o `tag` se já existir (o tag é o apelido que casa com o `source` do informe).

Uso:
  python fetch_escriturador.py --investimentos processed/brazil_investments.xlsx [--out memory/escriturador_memory.md]

Requires: openpyxl (stdlib para o resto: urllib, base64, json).
"""
import argparse, base64, json, os, re, sys, unicodedata, urllib.request
from pathlib import Path

B3 = "https://sistemaswebb3-listados.b3.com.br"
HDRS = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}

# escriturador (MAIÚSCULO) -> tag minúsculo que costuma aparecer no `source` do informes.json
TAGS = [("BRADESCO", "bradesco"), ("BTG", "btg"), ("ITAU", "itau"), ("ITAÚ", "itau"),
        ("BANCO DO BRASIL", "bb"), ("BB ", "bb"), ("SANTANDER", "santander"),
        ("XP ", "xp"), ("CAIXA", "caixa"), ("BNY", "bny"), ("GENIAL", "genial")]


def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().strip()
def acronym(tk): return re.sub(r"\d+[A-Za-z]?$", "", str(tk)).strip().upper()  # ALZR11->ALZR, B3SA3->B3SA, ROXO34->ROXO, XPML11B->XPML
def tag_for(esc):
    u = esc.upper()
    for needle, t in TAGS:
        if needle in u: return t
    return (norm(esc).split() or ["?"])[0].lower()


def b3_get(proxy, call, payload, tries=3):
    b = base64.b64encode(json.dumps(payload).encode()).decode()
    url = f"{B3}/{proxy}/{call}/{b}"
    for attempt in range(tries):
        try:
            with urllib.request.urlopen(urllib.request.Request(url, headers=HDRS), timeout=20) as r:
                raw = r.read().decode("utf-8", "ignore").strip()
            if not raw or raw == "null":       # resposta vazia: a B3 às vezes devolve vazio transitório
                continue
            data = json.loads(raw)
            if isinstance(data, str):          # a B3 devolve JSON duplo-encodado (string contendo JSON)
                data = json.loads(data)
            return data
        except Exception as e:
            if attempt == tries - 1:
                print(f"  ! B3 {call} falhou: {e}", file=sys.stderr)
    return None


def escriturador_fii(cnpj14, ticker):
    root = re.sub(r"\D", "", cnpj14)[:8]
    for tf in (7, 34, 21, 20):
        d = b3_get("fundsProxy/fundsCall", "GetListedSupplementFunds",
                   {"cnpj": root, "identifierFund": acronym(ticker), "typeFund": tf})
        if d and d.get("ifd"): return d["ifd"].strip()
    return None


def escriturador_acao(ticker):
    d = b3_get("listedCompaniesProxy/CompanyCall", "GetListedSupplementCompany",
               {"issuingCompany": acronym(ticker), "language": "pt-br"})
    if isinstance(d, list) and d:
        return (d[0].get("hasCommom") or d[0].get("hasPreferred") or "").strip() or None
    return None


def load_bens(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames if "bens_e_direitos" in s.lower()), None)
    if ws is None: sys.exit("IRPF_bens_e_direitos não encontrada no workbook")
    hdr = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        get = lambda k: r[hdr[k]] if k in hdr and hdr[k] < len(r) else None
        tk = norm(get("ticker"))
        if not tk: continue
        try: g, c = int(float(get("grupo"))), int(float(get("codigo")))
        except (TypeError, ValueError): g, c = None, None
        rows.append({"ticker": tk, "grupo": g, "codigo": c, "cnpj": norm(get("cnpj"))})
    return rows


def read_existing_tags(path):
    """preserva o tag já preenchido à mão (apelido que casa com o source do informe)."""
    tags = {}
    if os.path.exists(path):
        for ln in Path(path).read_text(encoding="utf-8").splitlines():
            s = ln.strip()
            if s.startswith("|") and s.count("|") >= 5:
                cells = [c.strip() for c in s.strip("|").split("|")]
                if len(cells) >= 4 and cells[0].upper() not in ("TICKER", "") and not set("".join(cells)) <= set("-: "):
                    tags[cells[0].upper()] = cells[3]
    return tags


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--investimentos", required=True)
    ap.add_argument("--out", default=os.path.join("memory", "escriturador_memory.md"))
    a = ap.parse_args()

    bens = load_bens(a.investimentos)
    prior = read_existing_tags(a.out)
    out_rows, n_ok = [], 0
    for it in bens:
        tk, g, c = it["ticker"], it["grupo"], it["codigo"]
        is_fii = g == 7
        is_eq = g == 3 or (g == 4 and c == 4)
        if not (is_fii or is_eq):
            continue                                            # Tesouro/RF/contas: sem escriturador de papel
        esc = escriturador_fii(it["cnpj"], tk) if is_fii else escriturador_acao(tk)
        if esc:
            n_ok += 1
            tag = prior.get(tk.upper()) or tag_for(esc)
            src = "B3 GetListedSupplementFunds (ifd)" if is_fii else "B3 GetListedSupplementCompany (hasCommom)"
            out_rows.append((tk, esc, it["cnpj"], tag, src))
            print(f"  {tk:8} -> {esc}  [tag {tag}]")
        else:
            out_rows.append((tk, "", it["cnpj"], prior.get(tk.upper(), ""), "B3 sem retorno — verificar manualmente"))
            print(f"  {tk:8} -> (B3 sem retorno)")

    Path(a.out).parent.mkdir(parents=True, exist_ok=True)
    L = ["# escriturador_memory (gerado por fetch_escriturador.py — fonte: API pública da B3)",
         "",
         "tag = apelido que aparece no `source` do rendimento no informes.json (preservado se já preenchido).",
         "Escriturador = campo `ifd` (FII/FI-Infra) / `hasCommom` (ação/BDR) da B3 — o mesmo de "
         "\"Informações Gerais do Fundo > Escriturador\".",
         "",
         "| ticker | escriturador | cnpj | tag | source |",
         "|---|---|---|---|---|"]
    for tk, esc, cnpj, tag, src in out_rows:
        L.append(f"| {tk} | {esc} | {cnpj} | {tag} | {src} |")
    Path(a.out).write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"\nOK -> {a.out}  ({n_ok}/{len(out_rows)} escrituradores resolvidos pela B3)")


if __name__ == "__main__":
    main()
