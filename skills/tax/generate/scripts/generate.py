#!/usr/bin/env python3
"""/generate — merge the B3 workbook + the unified informes transcription into the IRPF sections.

Produces a single ready-to-type workbook, irpf_consolidated.xlsx, one sheet per IRPF section:
  bens_e_direitos
  isentos                 (Rendimentos Isentos e Não Tributáveis)
  exclusiva_definitiva    (Tributação Exclusiva/Definitiva)

This is the deterministic second half of the old /consolidate, split into /read (the agent reads
every file in resources/ → processed/informes.json) and /generate (this script). The SAME unified
informes.json is consumed by /completeness, so there is one transcription, not two.

THE RULE that makes Bens e Direitos right
  For B3-custodied assets the DECLARED VALUE is the acquisition cost (preço médio × quantidade),
  taken from brazil_investments.xlsx (the b3 skill output: sheet IRPF_bens_e_direitos) — NEVER the
  informe's market value. But the IRPF GRUPO / CÓDIGO and the CNPJ come from the INFORME (the
  official document): e.g. NUIF11 is a FI-Infra incentivado (07/10), not a plain FII (07/03);
  only the informe carries the right código. Non-B3 assets (foreign accounts, bank funds,
  currency, US stocks) live ONLY in the informes and are taken verbatim from the JSON.

  Rendimentos (isentos / exclusiva) ALL come from the informes — never investimentos.

Inputs
  brazil_investments.xlsx   b3 skill output. Only IRPF_bens_e_direitos is read (preço médio, qtd, valor).
  informes.json             the unified transcription /read produces from every file in resources/.

Usage
  python generate.py --investimentos processed/brazil_investments.xlsx --json processed/informes.json [--outdir .]
  python generate.py --template processed/informes.json      # write an empty unified JSON skeleton

informes.json schema (one object, every list optional). `key` is the asset id (UPPERCASE for B3,
matching the b3 workbook's ticker); `source` is the PDF the value came from (tracking):
{
  "bens": [
    # B3 asset: value comes from brazil_investments.xlsx by key; grupo/codigo/cnpj from the informe
    {"key":"NUIF11","b3":true,"grupo":7,"codigo":10,"cnpj":"40963403000150","quantidade":9,
     "valor_2025":979.20,"source":"btg informe","discriminacao":"<optional override>"},
    # non-B3 asset: everything (incl. value) from the informe
    {"key":"NU-RDB","b3":false,"grupo":4,"codigo":2,"localizacao":105,"descr":"Saldo em RDB ...",
     "cnpj":"30680829000143","valor_2025":88547.65,"valor_2024":0.0,"source":"nubank informe"}
  ],
  "isentos":   [{"codigo":9, "key":"BBSE3","beneficiario":"Titular","cnpj":"...","fonte_pagadora":"...","descr":"...","valor_2025":0.0,"source":"..."}],
  "exclusiva": [{"codigo":10,"key":"VALE3","beneficiario":"Titular","cnpj":"...","descr":"JCP ...","valor_2025":0.0,"source":"..."}]
}

Requires: pandas, openpyxl
"""
import argparse, json, re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PURPLE, LPURPLE = "FF674EA7", "FFD9D2E9"
QTY_RE = re.compile(r"([\d.]+)\s*UNIDADES", re.I)
PM_RE  = re.compile(r"CUSTO\s+MEDIO[:\s]*R?\$?\s*([\d.]+,\d{2})", re.I)

# column order hints per section (known IRPF columns first; unknown keys appended)
COL_ORDER = {
    "bens":      ["origem","grupo","codigo","localizacao","cnpj","discriminacao","valor_2024","valor_2025","fonte"],
    "isentos":   ["codigo","key","beneficiario","cnpj","fonte_pagadora","descr","valor_2024","valor_2025","fonte"],
    "exclusiva": ["codigo","key","beneficiario","cnpj","descr","valor_2024","valor_2025","fonte"],
}
# columns to sum into a TOTAL row
def is_value_col(c): return c.startswith("valor")


def brnum(s):
    """'15.268,60' / 1234.56 / None -> float | None."""
    if s is None: return None
    if isinstance(s, (int, float)): return float(s)
    t = str(s).strip().replace("R$", "").strip()
    if not t: return None
    t = t.replace(".", "").replace(",", ".") if "," in t else t
    try: return float(t)
    except ValueError: return None


def fmt_cnpj(s):
    """14 raw digits -> 'XX.XXX.XXX/XXXX-XX' (kept as text so leading zeros survive). Else verbatim."""
    if s is None: return ""
    d = re.sub(r"\D", "", str(s))
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return str(s)


def as_int(v):
    """IRPF grupo/código as a clean int (3 not 3.0); pass through non-numerics."""
    try:
        f = float(v)
        return int(f) if f == int(f) else v
    except (TypeError, ValueError):
        return v


def load_investimentos(path):
    """ticker(UPPER) -> {valor_2024, valor_2025, quantidade, preco_medio, discriminacao}.
    Reads the b3 renda-variável Bens e Direitos sheet (only ações/FII/BDR — renda fixa is NOT here,
    its value comes from the informe). The sheet name varies (irpf_bens_e_direitos_renda_variavel),
    so match by substring."""
    xl = pd.ExcelFile(path)
    sheet = next((s for s in xl.sheet_names if "bens_e_direitos" in s.lower()), None)
    if sheet is None:
        raise SystemExit(f"sheet de bens (renda variável) não encontrada em {path}: {xl.sheet_names}")
    df = pd.read_excel(path, sheet_name=sheet)
    out = {}
    for _, r in df.iterrows():
        tk = str(r["ticker"]).strip()
        disc = str(r.get("discriminacao", "") or "")
        q = QTY_RE.search(disc); pm = PM_RE.search(disc)
        out[tk.upper()] = {
            "ticker": tk,
            "valor_2024": None if pd.isna(r.get("valor_2024")) else float(r["valor_2024"]),
            "valor_2025": None if pd.isna(r.get("valor_2025")) else float(r["valor_2025"]),
            "quantidade": float(q.group(1).replace(".", "")) if q else None,
            "preco_medio": brnum(pm.group(1)) if pm else None,
            "discriminacao": disc,
        }
    return out


def build_bens(items, inv, audit):
    """Merge: value (pm×qtd) from investimentos; grupo/codigo/cnpj/discriminação from the informe."""
    rows, seen = [], set()
    for it in items or []:
        is_b3 = bool(it.get("b3"))
        key = str(it.get("key", it.get("ticker", ""))).strip()
        fonte = it.get("source", it.get("fonte", ""))
        if is_b3:
            tk = key.upper()
            seen.add(tk)
            src = inv.get(tk)
            if not src:
                audit.append(f"  B3 '{tk}': no informe-classified ticker found in brazil_investments.xlsx — check the key")
                v24 = it.get("valor_2024"); v25 = it.get("valor_2025")
            else:
                v24, v25 = src["valor_2024"], src["valor_2025"]
            disc = it.get("discriminacao") or (src["discriminacao"] if src else f"{tk}")
            rows.append({"origem": "B3", "grupo": as_int(it.get("grupo")), "codigo": as_int(it.get("codigo")),
                         "localizacao": as_int(it.get("localizacao", 105)), "cnpj": fmt_cnpj(it.get("cnpj", "")),
                         "discriminacao": disc, "valor_2024": brnum(v24), "valor_2025": brnum(v25),
                         "fonte": fonte})
        else:
            rows.append({"origem": "informe", "grupo": as_int(it.get("grupo")), "codigo": as_int(it.get("codigo")),
                         "localizacao": as_int(it.get("localizacao", 105)),
                         "cnpj": fmt_cnpj(it.get("cnpj", it.get("pais", ""))),
                         "discriminacao": it.get("descr") or it.get("discriminacao", ""),
                         "valor_2024": brnum(it.get("valor_2024")), "valor_2025": brnum(it.get("valor_2025")),
                         "fonte": fonte})
    # AUDIT: B3 assets the b3 workbook holds but the informes-JSON forgot to classify
    for tk, s in inv.items():
        if tk not in seen and (s["valor_2025"] or s["valor_2024"]):
            audit.append(f"  '{s['ticker']}' is in brazil_investments.xlsx but NOT in informes.json "
                         f"(value R$ {s['valor_2025']:,.2f}) — add it with its informe grupo/codigo/cnpj")
    return rows


def norm_rows(items):
    """Rendimentos: pass-through, normalizing values/cnpj/codigo and mapping source -> fonte (column)."""
    out = []
    for it in items or []:
        r = {}
        for k, v in it.items():
            if k == "source":       r["fonte"] = v
            elif is_value_col(k):    r[k] = brnum(v)
            elif k == "cnpj":        r[k] = fmt_cnpj(v)
            elif k == "codigo":      r[k] = as_int(v)
            else:                    r[k] = v
        out.append(r)
    return out


def write_sheet(wb, title, rows, order_key, drop_first=False):
    ws = wb.active if not wb.sheetnames or (len(wb.sheetnames) == 1 and not drop_first and wb.active.max_row == 1 and wb.active.title == "Sheet") else wb.create_sheet(title)
    ws.title = title
    if not rows:
        ws.append(["(vazio)"]); return ws
    known = COL_ORDER.get(order_key, [])
    cols = [c for c in known if any(c in r for r in rows)]
    seen = set(cols)
    for r in rows:                       # append any extra keys once, in first-seen order
        for c in r:
            if c not in seen:
                seen.add(c); cols.append(c)
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    # TOTAL row
    vcols = [c for c in cols if is_value_col(c)]
    if vcols:
        total = ["TOTAL"] + [None] * (len(cols) - 1)
        for c in vcols:
            s = sum(r[c] for r in rows if isinstance(r.get(c), (int, float)))
            total[cols.index(c)] = round(s, 2)
        ws.append(total)
    style(ws, cols, vcols)
    return ws


def style(ws, cols, vcols):
    hdr = PatternFill("solid", fgColor=PURPLE)
    for cell in ws[1]:
        cell.fill = hdr; cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    last = ws.max_row
    if ws.cell(last, 1).value == "TOTAL":
        tf = PatternFill("solid", fgColor=LPURPLE)
        for cell in ws[last]:
            cell.fill = tf; cell.font = Font(bold=True)
    for c in vcols:
        L = get_column_letter(cols.index(c) + 1)
        for row in range(2, ws.max_row + 1):
            ws[f"{L}{row}"].number_format = "#,##0.00"
    for i, c in enumerate(cols, 1):
        w = max(len(str(c)), *(len(str(ws.cell(r, i).value or "")) for r in range(2, ws.max_row + 1)))
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--investimentos")
    ap.add_argument("--json")
    ap.add_argument("--outdir", default=".")
    ap.add_argument("--template", metavar="PATH", help="write an empty unified informes.json skeleton and exit")
    a = ap.parse_args()

    if a.template:
        skeleton = {
            "bens": [
                {"key": "EXEMPLO11", "b3": True, "grupo": 7, "codigo": 3, "cnpj": "", "quantidade": 0, "source": ""},
                {"key": "EXEMPLO-CONTA", "b3": False, "grupo": 6, "codigo": 1, "localizacao": 105,
                 "descr": "", "cnpj": "", "valor_2025": 0.0, "valor_2024": 0.0, "source": ""},
            ],
            "isentos":   [{"codigo": 9,  "key": "", "beneficiario": "Titular", "cnpj": "", "descr": "", "valor_2025": 0.0, "source": ""}],
            "exclusiva": [{"codigo": 10, "key": "", "beneficiario": "Titular", "cnpj": "", "descr": "", "valor_2025": 0.0, "source": ""}],
        }
        Path(a.template).write_text(json.dumps(skeleton, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"template -> {a.template}"); return

    if not (a.investimentos and a.json):
        ap.error("--investimentos and --json are required (or use --template)")

    inv = load_investimentos(a.investimentos)
    data = json.loads(Path(a.json).read_text(encoding="utf-8"))
    outdir = Path(a.outdir); outdir.mkdir(parents=True, exist_ok=True)
    audit = []

    # Single workbook, one sheet per IRPF section
    bd  = build_bens(data.get("bens"), inv, audit)
    iso = norm_rows(data.get("isentos"))
    exc = norm_rows(data.get("exclusiva"))

    wb = Workbook()
    write_sheet(wb, "bens_e_direitos", bd, "bens")
    write_sheet(wb, "isentos", iso, "isentos", drop_first=True)
    write_sheet(wb, "exclusiva_definitiva", exc, "exclusiva", drop_first=True)
    p = outdir / "irpf_consolidated.xlsx"; wb.save(p)

    def tot(rows, col="valor_2025"):
        return sum(r[col] for r in rows if isinstance(r.get(col), (int, float)))
    print(f"OK -> {p.name}")
    print(f"  bens_e_direitos       {len(bd)} itens, total 2025 R$ {tot(bd):,.2f}")
    print(f"  isentos               {len(iso)} itens, total R$ {tot(iso):,.2f}")
    print(f"  exclusiva_definitiva  {len(exc)} itens")
    print("\nAUDIT (B3 cross-check brazil_investments.xlsx × informes.json):")
    print("\n".join(audit) if audit else "  all B3 assets in brazil_investments.xlsx are classified in informes.json — OK")


if __name__ == "__main__":
    main()
